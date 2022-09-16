# -*- coding: utf-8 -*-
"""
Created on Wed Sep 12 14:09:08 2018

@author: nahuel.a.rios
"""

from pandas import ExcelFile,DataFrame,to_numeric,to_datetime,DateOffset,read_csv,concat,isnull,Series
import pandas as pd
from tkinter.ttk import Progressbar
from pyxlsb import open_workbook as open_xlsb
from os import environ,path,makedirs
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import pyperclip
from glob import glob
from shutil import move,copyfile
import numpy as np
from tkinter import Tk,Frame,N,W,E,S,StringVar,OptionMenu,Label,messagebox,Listbox,Scrollbar,Button,IntVar,Checkbutton,filedialog
from datetime import date,timedelta,datetime
from time import sleep
from math import ceil
from openpyxl import Workbook, drawing
from openpyxl.styles import Color, PatternFill, Font, Alignment, colors
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import input_parameters as ip
import chromedriver_autoinstaller
#import calendar


class BaseData():
    def __init__(self):
        print('Loading Base Data ' + datetime.now().time().strftime('%H:%M:%S'))
        self.root = Tk()        
        self.version = 'v1.31'
        self._datapath = '.\\BaseData'
        
        self.root.title('SRR Tool ' + self.version)
        self.estado = Label(self.root,text='Opening SRR Tool ' + self.version)
        self.estado.pack()
        self.increment = 0
        self.pbar = Progressbar(self.root, length=400)
        self.pbar.pack(padx=5, pady=5)
        self.root.after(500, self.advance)
        self.root.mainloop()
        
    def advance(self):
        self.pbar.step(5)
        self.increment += 5
        if self.increment == 20:
            self.chromeDriver()
            self.root.after(100, self.advance)
        elif self.increment == 40:
            self.loadTables()
            self.root.after(100,self.advance)
        elif self.increment < 100:
            self.root.after(100, self.advance) 
        else:
            self.root.destroy()

    def chromeDriver(self):        
        chdr = chromedriver_autoinstaller.install(cwd=True)  
        self.estado.config(text='Creating Chrome Web Driver')
        self.estado.update()
        try:
            usuario = environ.get('USERNAME')
            chrome_options = Options()
            prefs = {"profile.default_content_setting_values.notifications" : 2}
            chrome_options.add_experimental_option("prefs",prefs)
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("user-data-dir=C:\\Users\\"+usuario+"\\AppData\\Local\\Google\\Chrome\\User Data\\Profile 1")
            self.driver = webdriver.Chrome(chdr,options=chrome_options)
        except:
            self.driver = webdriver.Chrome(chdr)
        return
    
    def loadExcel(self,path,sheet,row): #SE USA PARA CARGAR LOS XLSX
        xlFile = ExcelFile(path)
        df = xlFile.parse(sheet_name=sheet,skiprows=row)
        return df
    
    def connectionIssue(self,link,nombre): #SI NO PUEDE CARGAR DIRECTAMENTE LOS ARCHIVOS LOS DESCARGA PARA INPUT MANUNAL
        messagebox.showerror(title='Connection Issue',message='Unable to connect to ' + nombre + '.\nFile will be automatically downloaded to your Downloads folder, please upload in next step')
        self.driver.get(link)
        path = filedialog.askopenfilename()
        return path

    def cargarITNAT(self): #COMO ITNAT ES XLSB TIENE SU FUNCION APARTE
        self.driver.get('https://ts.accenture.com/sites/PMOperations/PMTechPortal/PM Tech Global Documents/PM Gateway/Rules Engine/PMG Rules Engine user friendly data.xlsb')
        messagebox.showinfo(title='PMG Rules Engine',message='PMG Rules Engine Downloaded to your Downloads folder.\nPlease open file, remove restrictions and save. You will be requested to upload the file in the next step.')
        filename = filedialog.askopenfilename()
        filename = filename.replace('/','\\\\')
        df = []
        with open_xlsb(filename) as wb:
            with wb.get_sheet('Tax_Compliance Thresholds') as sheet:
                for row in sheet.rows():
                    df.append([item.v for item in row])
        df = DataFrame(df[2:], columns=df[1])
        return df

    def loadTables(self):
        self.estado.config(text='Loading SRR Tables')
        self.estado.update()
        #CARGA TODAS LAS HOJAS DEL SRR TABLES
        try:
            SRR_Tables = '//ts.accenture.com/sites/PMOperations3/PeopleMobilityServiceDelivery/Cross Tax/Shared Documents/Cross AEE/Tax Planning Tool/SRR Tables Python.xlsx'
            df_ToolVersion = self.loadExcel(SRR_Tables,'Tool Version',0)
        except:
            SRR_Tables = 'https://ts.accenture.com/sites/PMOperations3/PeopleMobilityServiceDelivery/Cross Tax/Shared Documents/Cross AEE/Tax Planning Tool/SRR Tables Python.xlsx'
            SRR_Tables = self.connectionIssue(SRR_Tables,'SRR Tables')
            df_ToolVersion = self.loadExcel(SRR_Tables,'Tool Version',0)         
        
        #VERIFICAR VERSION
        self.estado.config(text='Validating Tool Version')
        self.estado.update()
        self.pbar.step(10)
        self.increment +=10
        
        versionTool = df_ToolVersion['Release Number'].iloc[-1:].reset_index(drop=True)
        if self.version == versionTool[0]:
            self.estado.config(text='Loading SRR Tables')
            self.estado.update()
            self.pbar.step(10)
            self.increment +=10
            #SI LA VERSION ES VALIDA QUE CONTINUE CARGANDO, SINO QUE SALGA
            self.df_CountryNames = self.loadExcel(SRR_Tables,'CountryNames',0)
            self.df_CountryList = self.loadExcel(SRR_Tables,'CountryList',0)
            self.df_CountryList['Tax Year Start Date'] = self.df_CountryList['Tax Year Start Date'].dt.strftime('%m.%B-%d')
            self.df_CountryList.dropna(inplace=True)
            self.df_Exceptions = self.loadExcel(SRR_Tables,'Exceptions',0)
            self.df_TNATNames = self.loadExcel(SRR_Tables,'TNAT Names',0)
            self.df_ARE = self.loadExcel(SRR_Tables,'ARE',0)
            self.list_ARE = self.df_ARE['Company Code'].drop_duplicates().tolist()

            #CARGA TCDL TOOL
            self.estado.config(text='Loading TCDL Tool')
            self.estado.update()
            self.pbar.step(20)
            self.increment +=20
            try:
                TCDL_Tool = '//ts.accenture.com/sites/PMOperations3/PeopleMobilityServiceDelivery/Cross Tax/Shared Documents/Cross AEE/Tax Planning Tool/TCDL Tool.xlsx'
                self.df_TCDLTool = self.loadExcel(TCDL_Tool,'TC Criteria',0)
            except:
                TCDL_Tool = 'https://ts.accenture.com/sites/PMOperations3/PeopleMobilityServiceDelivery/Cross Tax/Shared Documents/Cross AEE/Tax Planning Tool/TCDL Tool.xlsx'
                TCDL_Tool = self.connectionIssue(TCDL_Tool,'TCDL Tool')
                self.df_TCDLTool = self.loadExcel(TCDL_Tool,'TC Criteria',0)
            
            #CARGA ITNAT
            self.estado.config(text='Loading iTNAT')
            self.estado.update()
            self.pbar.step(10)
            self.increment +=10
            
            df_iTNAT = self.cargarITNAT()
            #ACTUALIZA LOS NOMBRES DE ITNAT
            for _, row in self.df_TNATNames.iterrows():
                df_iTNAT.loc[df_iTNAT['Home Country']==row['TNAT Name'],'Home Country']=row['PMG Name']
                df_iTNAT.loc[df_iTNAT['Host Country']==row['TNAT Name'],'Host Country']=row['PMG Name']
                self.df_iTNAT = df_iTNAT[['Home Country','Host Country',
                               'Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered',
                               'Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']]                        
        else:
            messagebox.showerror(title='Incorrect Tool Version',message='Please download new tool (' + self.version +')')
            #LINK A LA NUEVA TOOL
            self.driver.get("https://myoffice.accenture.com/personal/h_fernandez_muriano_accenture_com/_layouts/15/onedrive.aspx?id=%2Fpersonal%2Fh_fernandez_muriano_accenture_com%2FDocuments%2FSRR%20Python")

class IE_Reports(object):
    def __init__(self,driver,reportByGU,hostCountries,hostGU,tyStart,reportStart,reportEnd,taxInfoYear,suppDoc,srrType):
        
    
#        if str(reportStart)[5:-9] != "01-01":   
#            yearpos = str(reportEnd)[:4]
#            if calendar.isleap(int(yearpos)) == True:
#               dia  = timedelta(days=1)
#               reportEnd = reportEnd + dia 
              
        self.driver = driver
        self._wait = WebDriverWait(self.driver, 60)
        self.reportByGU = reportByGU
        self.hostCountries = hostCountries
        self.hostGU = hostGU
        self.tyStart = tyStart
        self.reportStart = reportStart
        self.reportEnd = reportEnd
        self.taxInfoYear = taxInfoYear
        self.suppDoc = suppDoc
        self.srrType = srrType
        
        self.df_Travel_Plan = self.SRR_Report()
        
    def SRR_Report(self):
        print('Running SRR Report ' + datetime.now().time().strftime('%H:%M:%S'))
        def Run_Report(trStartDate,trEndDate,RNumber):
            def TravelStatus(strStatus):
                pos = filtro.index(strStatus) + 1 #OBTENGO LA POSICION EN LA LISTA DE LA GU ELEGIDA
                if pos < 10: 
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl25_divDropDown_ctl0' + str(pos)))); driver.execute_script("arguments[0].click();", option)
                else: 
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl25_divDropDown_ctl' + str(pos)))); driver.execute_script("arguments[0].click();", option)
                
            driver = self.driver
            wait = self._wait
            usuario = environ.get('USERNAME')
            
            #ACCEDER AL REPORTE
            driver.get('https://bi5.accenture.com/Reports/Pages/Report.aspx?ItemPath=%2f8044_People+Mobility+Gateway%2fTax%2fStrategic+Repatriation+Report')
            
            sleep(6)
            driver.switch_to_frame(wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div/paginated-report-viewer/div/iframe'))))
            
            #START DATE OPERATOR
            filtro = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl03_ddValue')))
            filtro.find_element_by_xpath('//*[@id="ReportViewerControl_ctl04_ctl03_ddValue"]/option[5]').click()
            driver.execute_script("arguments[0].click();", filtro)
            
            #TRAVEL START DATE
            fecha = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl05_cbNull'))); driver.execute_script("arguments[0].click();", fecha)
            fecha = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl05_txtValue')))
            fecha.send_keys(trStartDate.strftime('%m/%d/%Y'))
            #print(trStartDate)
            #END DATE OPERATOR
            filtro = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl11_ddValue')))
            filtro.find_element_by_xpath('//*[@id="ReportViewerControl_ctl04_ctl11_ddValue"]/option[3]').click()
            driver.execute_script("arguments[0].click();", filtro)
            #TRAVEL END DATE
            fecha = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl13_cbNull'))); driver.execute_script("arguments[0].click();", fecha)
            fecha = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl13_txtValue')))
            fecha.send_keys(trEndDate.strftime('%m/%d/%Y'))        
    
            'TAX INFORMATION YEAR'
            button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl35_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) #ABRO EL DESPLEGABLE
            filtro = driver.find_element_by_id('ReportViewerControl_ctl04_ctl35_divDropDown_ctl00');driver.execute_script("arguments[0].click();", filtro) #SELECCIONO ALL
            driver.execute_script("arguments[0].click();", button) #CIERRO EL DESPLEGABLE
            
            if self.srrType!='Outbound':
                'HOME GU'
                button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl29_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) #ABRO EL DESPLEGABLE
                filtro = driver.find_element_by_id('ReportViewerControl_ctl04_ctl29_divDropDown_ctl00');driver.execute_script("arguments[0].click();", filtro) #SELECCIONO ALL
                driver.execute_script("arguments[0].click();", button) #CIERRO EL DESPLEGABLE   
                
                'GU'
                button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) #ABRO EL DESPLEGABLE
                filtro = wait.until(EC.presence_of_element_located((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown'))) #SELECCIONO EL CAMPO DE LAS CAJAS
                filtro = filtro.text.splitlines() #CONVERTIMOS LA TABLA A UNA LISTA
                pos = filtro.index(self.hostGU) + 1 #OBTENGO LA POSICION EN LA LISTA DE LA GU ELEGIDA
                Filt_MU = self.hostGU
                if Filt_MU == 'United States':  
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown_ctl14'))); driver.execute_script("arguments[0].click();", option)
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown_ctl16'))); driver.execute_script("arguments[0].click();", option)
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown_ctl18'))); driver.execute_script("arguments[0].click();", option)
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown_ctl20'))); driver.execute_script("arguments[0].click();", option)
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown_ctl21'))); driver.execute_script("arguments[0].click();", option)
                else:
                    if pos < 10:
                        option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown_ctl0' + str(pos)))); driver.execute_script("arguments[0].click();", option)
                    else: 
                        option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_divDropDown_ctl' + str(pos)))); driver.execute_script("arguments[0].click();", option)
                driver.execute_script("arguments[0].click();", button) #CIERRO EL DESPLEGABLE

                'CORRER POR PAIS'
                if not self.reportByGU:
                    paises = ", ".join(self.hostCountries)
                    texto = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl31_txtValue'))); texto.send_keys(paises)
                    
            else:
                'HOME GU'
                button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl29_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) #ABRO EL DESPLEGABLE
                filtro = wait.until(EC.presence_of_element_located((By.ID,'ReportViewerControl_ctl04_ctl29_divDropDown'))) #SELECCIONO EL CAMPO DE LAS CAJAS
                filtro = filtro.text.splitlines() #CONVERTIMOS LA TABLA A UNA LISTA
                pos = filtro.index(self.hostGU) + 1 #OBTENGO LA POSICION EN LA LISTA DE LA GU ELEGIDA
                if pos < 10:
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl29_divDropDown_ctl0' + str(pos)))); driver.execute_script("arguments[0].click();", option)
                else: 
                    option = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl29_divDropDown_ctl' + str(pos)))); driver.execute_script("arguments[0].click();", option)
                driver.execute_script("arguments[0].click();", button) #CIERRO EL DESPLEGABLE   
                
                'GU'
                button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl33_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) #ABRO EL DESPLEGABLE
                filtro = driver.find_element_by_id('ReportViewerControl_ctl04_ctl33_divDropDown_ctl00'); driver.execute_script("arguments[0].click();", filtro)
                driver.execute_script("arguments[0].click();", button) #CIERRO EL DESPLEGABLE

                'CORRER POR PAIS'
                if not self.reportByGU:
                    paises = ", ".join(self.hostCountries)
                    texto = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl27_txtValue'))); texto.send_keys(paises)
                    
            'TRAVEL PLAN STATUS'
            button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl25_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) #ABRO EL DESPLEGABLE
            filtro = wait.until(EC.presence_of_element_located((By.ID,'ReportViewerControl_ctl04_ctl25_divDropDown'))) #SELECCIONO EL CAMPO DE LAS CAJAS
            # tabla = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="ctl32_ctl04_ctl25_divDropDown"]/span/div[1]/span/table')))
            filtro = filtro.text.splitlines() #CONVERTIMOS LA TABLA A UNA LISTA
            TravelStatus('Confirmed'); TravelStatus('On Hold'); TravelStatus('Pending')  
            driver.execute_script("arguments[0].click();", button) #CIERRO EL DESPLEGABLE
    
            button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl00'))); driver.execute_script("arguments[0].click();", button) #VIEW REPORT
            
            list_of_files_old = glob('C:\\Users\\' + usuario + '\\Downloads\\Strategic Repatriation Report*.xlsx') #NOMBRE DEL ULTIMO ARCHIVO ANTES DE DESCARGAR
            
            Carga = False
            
            while Carga != True:
                element = driver.find_element_by_id("ReportViewerControl_ctl04_ctl00")
                Carga = element.is_enabled()
                sleep(2)
            
            element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl05_ctl04_ctl00_ButtonImg")));driver.execute_script("arguments[0].click();", element)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ReportViewerControl_ctl05_ctl04_ctl00_Menu"]/div[5]/a')));driver.execute_script("arguments[0].click();", element)
                       
            
            while True: #ESPERA HASTA QUE ESTE EL ARCHIVO NUEVO
                list_of_files_new = glob('C:\\Users\\' + usuario + '\\Downloads\\Strategic Repatriation Report*.xlsx')
                if len(list_of_files_old) < len(list_of_files_new):
                    new_file = max(list_of_files_new, key=path.getctime)
                    break 
            while True:#CARGA EL REPORTE
                try:
                    xlFile = ExcelFile(new_file)
                    sheets = xlFile.sheet_names
                    del sheets[0]
                    df = None
                    for sheet in sheets:
                        if df is None:
                            df = xlFile.parse(sheet_name=sheet)
                        else:
                            df = df.append(xlFile.parse(sheet_name=sheet))
                    break
                except:
                    pass     
                
            df = df[(df['Policy Type'] != "750")]
            df = df[~df['Enterprise ID'].isnull()]
            df = df[(df['Program Name/Transfer Type'] != "Localization")]
            df = df[(df['Program Name/Transfer Type'] != "Localization Phase")]
            df = df[(df['Program Name/Transfer Type'] != "Localization Phase")]
            df = df[(df['Policy Type'] != "740") & (df['Home Country/Location'] != df['Destination Country/Location'])]


            df.rename(columns={
                    'Home Market Unit':'Home Geographic Unit',
                    'Destination Market Unit':'Destination Geographic Unit'},inplace=True)
            
            US_Lista = ["Northeast", "Midwest", "West" ,"South" , "United States"]
            
            df.loc[(df['Destination Geographic Unit'].isin(US_Lista)),['Destination Geographic Unit']] = "United States"
            df.loc[(df['Home Geographic Unit'].isin(US_Lista)),['Home Geographic Unit']] = "United States"
    
            move(new_file,self.suppDoc + '\\Strategic Repatriation Report ' + RNumber + '.xlsx')
            return df
        
        # nMonths = 6
        # if 'USA' in self.hostCountries:
        #     df = None
        #     total = ceil((self.reportEnd.month + 12)/nMonths)
        #     for i in range(total):
        #         print('Report ' + str(i+1) + ' of ' + str(total) + ' ' + datetime.now().time().strftime('%H:%M:%S'))
        #         start = self.reportStart + DateOffset(months=nMonths*i)
        #         end = start + DateOffset(months=nMonths) - timedelta(days=1)
        #         if df is None:
        #             df = Run_Report(end,start,str(i+1))
        #         else:
        #             df = df.append(Run_Report(end,start,str(i+1)))
        # else:
        df = Run_Report(self.reportEnd,self.reportStart,'1')
        
        return df
    
    def LBD_Report(self,df_TP):
        print('Running LBD Report ' + datetime.now().time().strftime('%H:%M:%S'))
        driver = self.driver
        wait = self._wait
        
        driver.get('https://bi4.accenture.com/Reports/Pages/Report.aspx?ItemPath=%2f2700_MyTimeandExpenses%2fRMS%2fRMS+Location+By+Day+Report') #CARGO LA PAGINA DE LBD E INGRESO FECHAS

        nMax = 100
        usuario = environ.get('USERNAME')   
        df = None
        #AGARRA LOS PRIMERIOS N° O HASTA EL MAX DE EID
        list_EID = df_TP['Enterprise ID'].drop_duplicates().tolist()
        totalReports = ceil(len(list_EID)/nMax)
        if len(list_EID) < nMax:
            EID = list_EID
            s_EID = "\n".join(list_EID)
        else:
            EID = list_EID[:nMax]
            s_EID = "\n".join(list_EID[:nMax])
        list_EID = list(set(list_EID) - set(EID)) 
        
        sleep(6)
        
        driver.switch_to_frame(wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div/paginated-report-viewer/div/iframe'))))
        
        sleep(1)
        
        fecha = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl03_txtValue")))
        driver.execute_script("arguments[0].setAttribute('value', '" + str(self.reportStart.strftime('%m/%d/%Y')) +"')", fecha);

        sleep(1)
        
        fecha1 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl05_txtValue")))
        driver.execute_script("arguments[0].setAttribute('value', '" + str(self.reportEnd.strftime('%m/%d/%Y')) +"')", fecha1);


#        #ENTERPRISE ID
        pyperclip.copy(s_EID)
        button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl09_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button)        
        elem = driver.find_element_by_id('ReportViewerControl_ctl04_ctl09_divDropDown_ctl00')
        elem.send_keys(Keys.CONTROL, 'v') #paste        
        driver.execute_script("arguments[0].click();", button)  
#        #TYPE OF DAY
#        button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ct-l04_ctl17_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) 
#        elem = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl17_divDropDown_ctl00'))); driver.execute_script("arguments[0].click();", elem) 
#        driver.execute_script("arguments[0].click();", button) 
        
        #CLICK
        
        sleep(1)
        
        element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl00")));driver.execute_script("arguments[0].click();", element)
        
        
        ReportNumber = 1
        print('LBD ' + str(ReportNumber) + ' of ' + str(totalReports) + ' ' + datetime.now().time().strftime('%H:%M:%S'))
        while len(list_EID)!=0:
            #AGARRO SIGUIENTE TANDA
            if len(list_EID) < nMax:
                EID = list_EID
                s_EID = "\n".join(list_EID)
            else:
                EID = list_EID[:nMax]
                s_EID = "\n".join(list_EID[:nMax])
            list_EID = list(set(list_EID) - set(EID))
            list_of_files_old = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv') #NOMBRE DEL TODOS LOS ULTIMOS ARCHIVOS ANTES DE DESCARGAR       
            
            
            
            Carga = False
            
            while Carga != True:
                element = driver.find_element_by_id("ReportViewerControl_ctl04_ctl00")
                Carga = element.is_enabled()
                sleep(2)
            
            element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl05_ctl04_ctl00_ButtonImg")));driver.execute_script("arguments[0].click();", element)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ReportViewerControl_ctl05_ctl04_ctl00_Menu"]/div[2]/a')));driver.execute_script("arguments[0].click();", element)
                            

            while True: #NOMBRE DEL NUEVO ARCHIVO DESCARGADO
                list_of_files_new = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv')
                if len(list_of_files_new) > len(list_of_files_old):
                    break     
            new_file = list(set(list_of_files_new) - set(list_of_files_old))[0]          
            #CARGA EL ARCHIVO
            while True:
                try:
                    if df is None:
                        df = read_csv(new_file)
                        df[['PersonnelNbr','Year','Date','CompanyCd']] = df[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                    else:
                        df1 = read_csv(new_file)
                        df1[['PersonnelNbr','Year','Date','CompanyCd']] = df1[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                        df = df.append(df1)
                    break
                except:
                    pass
            move(new_file,self.suppDoc + '\RMS Location By Day Report ' + str(ReportNumber) + '.csv') #MUEVE ARCHIVO
            ReportNumber+=1
            #PONGO A DESCARGAR LA TANDA NUEVA
            
            
            
            #ENTERPRISE ID
#            pyperclip.copy(s_EID)
#            button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl09_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button)
#            driver.find_element_by_id('ReportViewerControl_ctl04_ctl09_divDropDown_ctl00').clear()
#            elem = driver.find_element_by_id('ReportViewerControl_ctl04_ctl09_divDropDown_ctl00')
#            elem.send_keys(Keys.CONTROL, 'v') #paste
            sleep(1)
            
            pyperclip.copy(s_EID)
            button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl09_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button)        
            elem = driver.find_element_by_id('ReportViewerControl_ctl04_ctl09_divDropDown_ctl00')
            elem.clear()
            elem.send_keys(Keys.CONTROL, 'v') #paste  
            driver.execute_script("arguments[0].click();", button)  

            sleep(1)
            element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl00")));driver.execute_script("arguments[0].click();", element)
                
            
            print('LBD ' + str(ReportNumber) + ' of ' + str(totalReports) + ' ' + datetime.now().time().strftime('%H:%M:%S'))

        'DESCARGA ULTIMA TANDA DE REPORTES'
        list_of_files_old = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv') #NOMBRE DEL TODOS LOS ULTIMOS ARCHIVOS ANTES DE DESCARGAR
        
        Carga = False
        
        while Carga != True:
            element = driver.find_element_by_id("ReportViewerControl_ctl04_ctl00")
            Carga = element.is_enabled()
            sleep(2)
        
        element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl05_ctl04_ctl00_ButtonImg")));driver.execute_script("arguments[0].click();", element)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ReportViewerControl_ctl05_ctl04_ctl00_Menu"]/div[2]/a')));driver.execute_script("arguments[0].click();", element)
                    
        
        #NOMBRE DEL NUEVO ARCHIVO DESCARGADO
        while True:
            list_of_files_new = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv')
            if len(list_of_files_new) > len(list_of_files_old):
                break
        new_file = list(set(list_of_files_new) - set(list_of_files_old))[0]
        #CARGA EL ARCHIVO
        while True:
            try:
                if df is None:
                    df = read_csv(new_file)
                    df[['PersonnelNbr','Year','Date','CompanyCd']] = df[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                else:
                    df1 = read_csv(new_file)
                    df1[['PersonnelNbr','Year','Date','CompanyCd']] = df1[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                    df = df.append(df1)
                break
            except:
                pass
        #MUEVE AL SUPPORTING DOC
        move(new_file,self.suppDoc + '\RMS Location By Day Report ' + str(ReportNumber) + '.csv') #MUEVE ARCHIVO
        ReportNumber+=1

        return df

    def LBD_Report_by_PN(self,df_PN):
        print('Running LBD Report by PN ' + datetime.now().time().strftime('%H:%M:%S'))
        driver = self.driver
        wait = self._wait
        
        driver.get('https://bi4.accenture.com/Reports/Pages/Report.aspx?ItemPath=%2f2700_MyTimeandExpenses%2fRMS%2fRMS+Location+By+Day+Report') #CARGO LA PAGINA DE LBD E INGRESO FECHAS

        nMax = 100
        usuario = environ.get('USERNAME')   
        df = None
        df_PN['PN'] = df_PN['PN'].astype(int,errors='ignore').astype(str)
        #AGARRA LOS PRIMERIOS N° O HASTA EL MAX DE EID
        list_EID = df_PN['PN'].tolist()
        list_EID = list(set(list_EID))
        totalReports = ceil(len(list_EID)/nMax)
        if len(list_EID) < nMax:
            EID = list_EID
            s_EID = "\n".join(list_EID)
        else:
            EID = list_EID[:nMax]
            s_EID = "\n".join(list_EID[:nMax])
        list_EID = list(set(list_EID) - set(EID)) 
        
        sleep(6)
        
        
        driver.switch_to_frame(wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div/paginated-report-viewer/div/iframe'))))
        
        sleep(1)
        
        fecha = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl03_txtValue")))
        driver.execute_script("arguments[0].setAttribute('value', '" + str(self.reportStart.strftime('%m/%d/%Y')) +"')", fecha);

        sleep(1)
        
        fecha1 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl05_txtValue")))
        driver.execute_script("arguments[0].setAttribute('value', '" + str(self.reportEnd.strftime('%m/%d/%Y')) +"')", fecha1);



        #PN
        sleep(1)

        pyperclip.copy(s_EID)
        button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl11_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button)
        elem = driver.find_element_by_id('ReportViewerControl_ctl04_ctl11_divDropDown_ctl00')
        elem.send_keys(Keys.CONTROL, 'v') #paste
        driver.execute_script("arguments[0].click();", button)
        
#        #TYPE OF DAY
#        button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl17_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button) 
#        elem = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl17_divDropDown_ctl00'))); driver.execute_script("arguments[0].click();", elem) 
#        driver.execute_script("arguments[0].click();", button) 
 
        sleep(1)

        element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl00")));driver.execute_script("arguments[0].click();", element)

        ReportNumber = 1
        print('LBD ' + str(ReportNumber) + ' of ' + str(totalReports) + ' ' + datetime.now().time().strftime('%H:%M:%S'))
        while len(list_EID)!=0:
            #AGARRO SIGUIENTE TANDA
            if len(list_EID) < nMax:
                EID = list_EID
                s_EID = "\n".join(list_EID)
            else:
                EID = list_EID[:nMax]
                s_EID = "\n".join(list_EID[:nMax])
            list_EID = list(set(list_EID) - set(EID))
            list_of_files_old = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv') #NOMBRE DEL TODOS LOS ULTIMOS ARCHIVOS ANTES DE DESCARGAR       
            

            Carga = False
            
            while Carga != True:
                element = driver.find_element_by_id("ReportViewerControl_ctl04_ctl00")
                Carga = element.is_enabled()
                sleep(2)
            
            element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl05_ctl04_ctl00_ButtonImg")));driver.execute_script("arguments[0].click();", element)
            element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ReportViewerControl_ctl05_ctl04_ctl00_Menu"]/div[2]/a')));driver.execute_script("arguments[0].click();", element)
                  
            
            while True: #NOMBRE DEL NUEVO ARCHIVO DESCARGADO
                list_of_files_new = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv')
                if len(list_of_files_new) > len(list_of_files_old):
                    break     
            new_file = list(set(list_of_files_new) - set(list_of_files_old))[0]          
            #CARGA EL ARCHIVO
            while True:
                try:
                    if df is None:
                        df = read_csv(new_file)
                        df[['PersonnelNbr','Year','Date','CompanyCd']] = df[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                    else:
                        df1 = read_csv(new_file)
                        df1[['PersonnelNbr','Year','Date','CompanyCd']] = df1[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                        df = df.append(df1)
                    break
                except:
                    pass
            move(new_file,self.suppDoc + '\RMS Location By Day Report by PN ' + str(ReportNumber) + '.csv') #MUEVE ARCHIVO
            ReportNumber+=1
            #PONGO A DESCARGAR LA TANDA NUEVA
            #PN
#            pyperclip.copy(s_EID)
#            button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl11_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button)
#            driver.find_element_by_id('ReportViewerControl_ctl04_ctl11_divDropDown_ctl00').clear()
#            elem = driver.find_element_by_id('ReportViewerControl_ctl04_ctl11_divDropDown_ctl00')
#            elem.send_keys(Keys.CONTROL, 'v') #paste
#            driver.execute_script("arguments[0].click();", button)
            
            sleep(1)
            
            pyperclip.copy(s_EID)
            button = wait.until(EC.element_to_be_clickable((By.ID,'ReportViewerControl_ctl04_ctl11_ddDropDownButton'))); driver.execute_script("arguments[0].click();", button)
            elem = driver.find_element_by_id('ReportViewerControl_ctl04_ctl11_divDropDown_ctl00')
            elem.clear()
            elem.send_keys(Keys.CONTROL, 'v') #paste
            driver.execute_script("arguments[0].click();", button)
            
            sleep(1)
            element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl04_ctl00")));driver.execute_script("arguments[0].click();", element)
       

            print('LBD ' + str(ReportNumber) + ' of ' + str(totalReports) + ' ' + datetime.now().time().strftime('%H:%M:%S'))
    
        'DESCARGA ULTIMA TANDA DE REPORTES'
        list_of_files_old = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv') #NOMBRE DEL TODOS LOS ULTIMOS ARCHIVOS ANTES DE DESCARGAR
        

        Carga = False
        
        while Carga != True:
            element = driver.find_element_by_id("ReportViewerControl_ctl04_ctl00")
            Carga = element.is_enabled()
            sleep(2)
        
        element = wait.until(EC.element_to_be_clickable((By.ID, "ReportViewerControl_ctl05_ctl04_ctl00_ButtonImg")));driver.execute_script("arguments[0].click();", element)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ReportViewerControl_ctl05_ctl04_ctl00_Menu"]/div[2]/a')));driver.execute_script("arguments[0].click();", element)
                      
        #NOMBRE DEL NUEVO ARCHIVO DESCARGADO
        while True:
            list_of_files_new = glob('C:\\Users\\' + usuario + '\\Downloads\\RMS Location By Day Report*.csv')
            if len(list_of_files_new) > len(list_of_files_old):
                break
        new_file = list(set(list_of_files_new) - set(list_of_files_old))[0]
        #CARGA EL ARCHIVO
        while True:
            try:
                if df is None:
                    df = read_csv(new_file)
                    df[['PersonnelNbr','Year','Date','CompanyCd']] = df[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                else:
                    df1 = read_csv(new_file)
                    df1[['PersonnelNbr','Year','Date','CompanyCd']] = df1[['PersonnelNbr','Year','Date','CompanyCd']].astype(int)
                    df = df.append(df1)
                break
            except:
                pass
        #MUEVE AL SUPPORTING DOC
        move(new_file,self.suppDoc + '\RMS Location By Day Report by PN ' + str(ReportNumber) + '.csv') #MUEVE ARCHIVO
        ReportNumber+=1
        return df        

def TCDL_Analysis(df,df_CountryList,df_TCDLTool,dt_TYStart,st_GU):
    print('Running TCDL Analysis ' + datetime.now().time().strftime('%H:%M:%S'))
    #REMUEVE LOS PAISES QUE TIENEN UN TY DISTINTO; APLICA CUANDO SE CORRE POR GU
    df_RemoveCountries = df_CountryList[(df_CountryList['Geographic Unit Description']==st_GU) & 
                                         (df_CountryList['Tax Year Start Date']!=dt_TYStart.strftime('%m.%B-%d'))]['Country Name']
    
    df = df.loc[~df['Destination Country'].isin(df_RemoveCountries)]
    
    df = df[(df['Policy Type'] == "710") | (df['Policy Type'] == "810") | (df['Policy Type'] == "1430") | (df['Policy Type'] == "740") | (df['Policy Type'] == "Other")] #DEJAMOS SOLO POLICY ELEGIBLES
    
    df['Assignment Length'] = ((df['Travel End Date']-df['Travel Start Date'])/np.timedelta64(1, 'D')).astype(int) + 1 #CALCULAR DURACION DE TRAVEL
    df['INCLUDE'] = False #COLUMNA INCLUYENDO DIAS DE LA TCDL TOOL
    df_TCDLTool = df_TCDLTool[(df_TCDLTool['Destination GU']=='All') | (df_TCDLTool['Destination GU']==st_GU)] #SE QUEDA CON LA GU DE ANALISIS DENTRO DEL TCDL
    df['Activity Type'] = Series(df['Activity Type'], dtype = object)
    
    for _, row in df_TCDLTool.iterrows(): #LOOP EN EL TCDL PARA DETERMINAR SI LA ASIGNACION DEBE SER INCLUIDA
        df.loc[(((df['Home Country']==row['Home Country']) | (row['Home Country']=='All')) &
               ((df['Destination Country']==row['Destination Country']) | (row['Destination Country']=='All')) &
               ((df['Destination Geographic Unit']==row['Destination GU']) | (row['Destination GU']=='All')) &
               ((df['Project Visa Type']==row['Project Visa Type']) | (row['Project Visa Type']=='All')) &
               ((df['Activity Type']==row['Activity Type']) | (row['Activity Type']=='All')) &
               ((df['Travel Assignment Category']==row['Travel Assignment Category']) | (row['Travel Assignment Category']=='All')) &
               (df['Assignment Length']>=row['Assignment Length']) &
               (df['Travel End Date']>=dt_TYStart)) | (df['Taxable End Date']>=dt_TYStart),
               'INCLUDE'] = True
                
    df['Code'] = df['People Key'].astype(int).astype(str) + df['Destination Country']
    PopInclude = df[df['INCLUDE']==True]['Code'].drop_duplicates() #OBTENEMOS PEOPLE KEY DE LA POBLACION A INCLUIR
    df.loc[df['Code'].isin(PopInclude), "INCLUDE"] = True #VERIFICAMOS DE INCLUIR A TODAS LAS ASIGNACIONES DE LOS EMPLEADOS QUE TIENEN UNA ELEGIBLE
    del(df['Code'])
    df = df[df['INCLUDE']==True]
    df.drop(columns=['Assignment Length','INCLUDE'],inplace=True)
    return df
       
def Tax_Threshold(df_iTNAT,df_SRR,df_Exceptions,srrType):
    print('Populating Tax Thresholds ' + datetime.now().time().strftime('%H:%M:%S'))
    df_SRR['Assignment Duration'] = ((df_SRR['Travel End Date']-df_SRR['Travel Start Date'])/np.timedelta64(1, 'D')).astype(int) + 1 #CALCULAR DURACION DE TRAVEL
    df_SRR['Assignment Duration'] = to_numeric(df_SRR['Assignment Duration'], errors='coerce')
    
    #MERGE DE ITNAT CON SRR
    df_iTNAT.set_index(['Home Country','Host Country'],inplace=True,drop=True)
    df_SRR = df_SRR.merge(df_iTNAT,how='left',left_on=['Home Country','Destination Country'],right_index=True)
    df_iTNAT.reset_index(inplace=True)
    #A LOS QUE SON NA, APLICA 90 ROLLING
    df_SRR[df_SRR.columns[len(df_SRR.columns)-2]] = to_numeric(df_SRR[df_SRR.columns[len(df_SRR.columns)-2]], errors='coerce')
    df_SRR.loc[df_SRR[df_SRR.columns[len(df_SRR.columns)-2]].isnull(),
                                     [df_SRR.columns[len(df_SRR.columns)-2],df_SRR.columns[len(df_SRR.columns)-1],'SRR Comments']]=[90,'Rolling','No Tax Threshold Information']
    df_SRR['Day Type']='PP'
    
    #EXCEPTIONS
    print('Tax Threshold Exceptions ' + datetime.now().time().strftime('%H:%M:%S'))
    if srrType == 'Outbound':
        ExceptionCountry = 'Home Country'
    else:
        ExceptionCountry = 'Destination Country'
    df_Exceptions = df_Exceptions[df_Exceptions['SRR Type']==srrType]
    df_Exceptions = df_Exceptions[df_Exceptions['Exception Country'].isin(df_SRR[ExceptionCountry].drop_duplicates())]
    df_Exceptions.sort_values(by='Order',ascending=False,inplace=True)
    df_Exceptions.reset_index(drop=True,inplace=True)
    if df_Exceptions is not None:
        df_Exceptions.fillna('N/A',inplace=True)
        df_Exceptions['Travel Start Date'] = to_datetime(df_Exceptions['Travel Start Date'], errors='coerce')
        df_Exceptions['Travel End Date'] = to_datetime(df_Exceptions['Travel End Date'], errors='coerce')
        df_Exceptions['Taxable Start Date'] = to_datetime(df_Exceptions['Taxable Start Date'], errors='coerce')
        df_Exceptions['Taxable End Date'] = to_datetime(df_Exceptions['Taxable End Date'], errors='coerce')
        df_Exceptions['Assignment Duration'] = to_numeric(df_Exceptions['Assignment Duration'], errors='coerce')
        for _, row in df_Exceptions.iterrows():
            df = df_SRR[df_SRR[ExceptionCountry]==row['Exception Country']]
            i=6
            while i <= len(df_Exceptions.columns)-1:
                header = df_Exceptions.columns[i]
                if i==6:
                    if isnull(row[header])==False:
                        df = df[df[header]>=row[header]]
                elif i==7 or i==9:
                    if isnull(row[header])==False:
                        df = df[df[header]<=row[header]]
                elif i==8 or i==10:
                    if isnull(row[header])==False:
                        df = df[df[header]>=row[header]]
                elif row[header]=='IsEmpty':
                    df = df[df[header].isnull()]
                elif row[header]!='N/A':
                    df = df[df[header].str.contains(row[header],na=False,flags=re.IGNORECASE)]
                i+=1
            df_SRR.loc[df_SRR['Travel Plan Number'].isin(df['Travel Plan Number']),
                       [df_SRR.columns[len(df_SRR.columns)-3],df_SRR.columns[len(df_SRR.columns)-2],df_SRR.columns[len(df_SRR.columns)-1],'SRR Comments']]=[row['Guideline'],row['Type'],row['Day Type'],'Tax Threshold Exception']
        df_SRR.drop(columns=['Assignment Duration'],inplace=True)
    return df_SRR

def Days_Analysis(df_SRR,df_LBD,df_CountryNames,reportStart,reportEnd,tyStart):
    print('Analyzing days ' + datetime.now().time().strftime('%H:%M:%S'))
    #TRIM
    df_LBD['HomeCountry'] = df_LBD['HomeCountry'].str.strip()
    df_LBD['Location1'] = df_LBD['Location1'].str.strip()
    
    #RENOMBRA A NOMBRE PMG  
    for _, row in df_CountryNames.iterrows():
        print("MyTE: " + row['MyTE Name']);print("PMG : " + row['PMG Name'])
        df_LBD.loc[df_LBD['HomeCountry']==row['MyTE Name'],'HomeCountry']=row['PMG Name']
        df_LBD.loc[df_LBD['Location1']==row['MyTE Name'],'Location1']=row['PMG Name']
        
    #CREA COLUMNA DE DATE VALUE
    # df_LBD.dropna(subset=['EnterpriseId'],inplace=True)
    df_LBD['Date'] = df_LBD['Date'].astype(int)
    df_LBD['Year'] = df_LBD['Year'].astype(int)
    df_LBD['DateVal'] = to_datetime(df_LBD['Date'].astype(str) + '-' + df_LBD['Month'] + '-' + df_LBD['Year'].astype(str),format='%d-%B-%Y', errors='coerce')
    
    #ANALISIS DE DOA
    print('DOA Analysis (' + str(len(df_SRR)) + ' assignments) '  + datetime.now().time().strftime('%H:%M:%S'))
    df = df_LBD[['People Key','Location1','DateVal']].merge(df_SRR[['People Key','Destination Country','Travel Start Date','Travel End Date']],how='inner',left_on=['People Key','Location1'],right_on=['People Key','Destination Country'])
    df['DOA'] = None
    df.loc[(df['DateVal']>=df['Travel Start Date']) & (df['DateVal']<=df['Travel End Date']), 'DOA'] = 'DAYS IN'
    df = df[df['DOA']=='DAYS IN'][['People Key','Location1','DateVal','DOA']].drop_duplicates()
    df_LBD = df_LBD.merge(df,how='left',left_on=['People Key','Location1','DateVal'],right_on=['People Key','Location1','DateVal'])
    df_LBD['DOA'].fillna('DAYS OUT',inplace=True)
    
    #CREA LA COLUMNA CON EL PERIODO
    df_LBD['Period'] = df_LBD['Month'] + '-' + df_LBD['Year'].astype(str)
    
    #NOS QUEDAMOS SOLO CON LAS COLUMNAS QUE NECESITAMOS
    df_LBD = df_LBD[['People Key','HomeCountry','Location1','TypeOfDay','DateVal','DOA','Period']]
    df_LBD.drop_duplicates(inplace=True)
    df_LBD.reset_index(drop=True,inplace=True)
    
    #REMUEVE LOS HOME = HOST (TRANSFERS)
    df_Location = df_LBD[(df_LBD['HomeCountry']!=df_LBD['Location1']) &
                         df_LBD['Location1'].isin(df_SRR['Destination Country'].drop_duplicates())]
    
    #PIVOT 1 CON WD Y OTRA PP
    print('Populating days ' + datetime.now().time().strftime('%H:%M:%S'))
    df_WD = DataFrame.pivot_table(df_Location[df_Location['TypeOfDay']=='Work Day'],values='DateVal',
                                  index=['People Key','Location1'],columns=['Period'],aggfunc=lambda x: len(x.unique())).fillna(0)
    df_PP = DataFrame.pivot_table(df_Location,values='DateVal',
                                  index=['People Key','Location1'],columns=['Period'],aggfunc=lambda x: len(x.unique())).fillna(0)
    
    #DA ORDEN A LAS COLUMNAS
    colDate = reportStart.replace(day=1)
    colName = []
    for i in range(24):
        colName.append(colDate.strftime('%B-%Y'))
        if colDate.strftime('%B-%Y') not in df_WD: df_WD[colDate.strftime('%B-%Y')] = 0
        if colDate.strftime('%B-%Y') not in df_PP: df_PP[colDate.strftime('%B-%Y')] = 0
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
    df_WD = df_WD[colName]; df_PP = df_PP[colName]
    
    '''
    POPULA LOS DIAS EN LOS TY
    1. SEPARA LA POBLACION QUE TIENE THRESHOLD WD Y POPULA CON TABLA WD
    2. SEPARA LA POBLACION QUE TIENE THRESHOLD PP Y POPULA CON TABLA PP
    3. VUELVE A JUNTAR LAS 2 TABLAS
    '''
    if len(df_Location) != 0:
        df_1 = df_SRR[df_SRR['Day Type']=='WD'].merge(df_WD,how='left',left_on=['People Key','Destination Country'],right_index=True)
        df_2 = df_SRR[df_SRR['Day Type']=='PP'].merge(df_PP,how='left',left_on=['People Key','Destination Country'],right_index=True)
        df_SRR = df_1.append(df_2)
        df_SRR[colName]=df_SRR[colName].fillna(0)
    else:
        for c in colName:
            df_SRR[c]=0
    
    #CALCULA EL MAXA
    print('Calculating MAXA ' + datetime.now().time().strftime('%H:%M:%S'))
    for i in range(12):
        df_SRR[str(i+1) + ' TY'] = 0
        df_SRR[str(i+1) + ' Rolling'] = 0
        #HACE 12 COLUMNAS CON DIAS EN TY
        for j in range(i+12,11,-1):
            df_SRR[str(i+1) + ' TY'] += df_SRR[colName[j]]
        #HACE 12 COLUMNAS CON DIAS ROLLING
        for j in range(i+1,i+13):
            df_SRR[str(i+1) + ' Rolling'] += df_SRR[colName[j]]      
        #CAMBIA A 0 EN LA COLUMNA ROLLING SI LA COLUMNA TY ES 0
        df_SRR.loc[df_SRR[str(i+1) + ' TY']==0,str(i+1) + ' Rolling'] = 0            
    #CALCULA EL MAXA EN ESOS 12 MESES
    df_SRR['MAXA'] = df_SRR[['1 Rolling','2 Rolling','3 Rolling','4 Rolling','5 Rolling','6 Rolling','7 Rolling','8 Rolling','9 Rolling','10 Rolling','11 Rolling','12 Rolling']].max(axis=1)
    
    ''' EXCEPTIONS EN EL MAXA
    1. AUSTRIA: 6 MESES ROLLING
    2. CZECH REPUBLIC Y SLOVAKIA CON THRESHOLD 16/90 ROLLING WD: 6 MONTH ROLLING * DEBE CONTAR SOLO LOS DIAS DE ENERO 2018 EN ADELANTE
    '''
    ####### EXCEPTION 6 MONTH ROLLING
    print('6 Month Exception Analysis ' + datetime.now().time().strftime('%H:%M:%S'))
    #CALCULA EL MAXA
    for i in range(12):
        df_SRR[str(i+1) + ' Rolling'] = 0
        df_SRR[str(i+1) + ' Rolling 2018'] = 0
        #HACE 12 COLUMNAS CON DIAS ROLLING 6 MESES
        for j in range(i+7,i+13):
            df_SRR[str(i+1) + ' Rolling'] += df_SRR[colName[j]]   
        #HACE 12 COLUMNAS CON DIAS ROLLING A PARTIR DE 2018
        for j in range(i+7,i+13):
            if int(colName[j][-4:]) >= 2018:
                df_SRR[str(i+1) + ' Rolling 2018'] += df_SRR[colName[j]]  
        #CAMBIA A 0 EN LA COLUMNA ROLLING SI LA COLUMNA TY ES 0
        df_SRR.loc[df_SRR[str(i+1) + ' TY']==0,str(i+1) + ' Rolling'] = 0 
        df_SRR.loc[df_SRR[str(i+1) + ' TY']==0,str(i+1) + ' Rolling 2018'] = 0            
    #CALCULA EL MAXA EN ESOS 12 MESES
    df_SRR.loc[df_SRR['Destination Country']=='Austria','MAXA'] = df_SRR[['1 Rolling','2 Rolling','3 Rolling','4 Rolling','5 Rolling','6 Rolling','7 Rolling','8 Rolling','9 Rolling','10 Rolling','11 Rolling','12 Rolling']].max(axis=1)
    df_SRR.loc[((df_SRR['Destination Country']=='Czech Republic') |
            (df_SRR['Destination Country']=='Slovakia')) &
        ((df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered']==16) |
                (df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered']==90)) &
                (df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling'),'MAXA'] = df_SRR[['1 Rolling 2018','2 Rolling 2018','3 Rolling 2018','4 Rolling 2018','5 Rolling 2018','6 Rolling 2018','7 Rolling 2018','8 Rolling 2018','9 Rolling 2018','10 Rolling 2018','11 Rolling 2018','12 Rolling 2018']].max(axis=1)
    #REMUEVE LAS COLUMNAS DE LOS 12 MESES
    df_SRR.drop(columns=['1 TY','2 TY','3 TY','4 TY','5 TY','6 TY','7 TY','8 TY','9 TY','10 TY','11 TY',
                         '1 Rolling','2 Rolling','3 Rolling','4 Rolling','5 Rolling','6 Rolling','7 Rolling','8 Rolling','9 Rolling','10 Rolling','11 Rolling','12 Rolling',
                         '1 Rolling 2018','2 Rolling 2018','3 Rolling 2018','4 Rolling 2018','5 Rolling 2018',
                         '6 Rolling 2018','7 Rolling 2018','8 Rolling 2018','9 Rolling 2018','10 Rolling 2018','11 Rolling 2018','12 Rolling 2018'],inplace=True)
    ################## END OF EXCEPTIONS
    
    #CALULA LOS DIAS EN LOS ULTIMOS 11 MESES, SE USA CUANDO ES CALENDAR Y POR EL MAXA NO LLEGA
    df_SRR['DAYS IN LAST 11'] = 0
    for i in range(colName.index(reportEnd.strftime('%B-%Y')),colName.index(reportEnd.strftime('%B-%Y'))-11,-1): 
        df_SRR['DAYS IN LAST 11'] += df_SRR[colName[i]]
    
    #CALULA LOS DIAS EN LOS ULTIMOS 12 MESES, SE USA CUANDO ES ROLLING Y POR EL MAXA NO LLEGA
    df_SRR['DAYS IN LAST 12'] = 0
    for i in range(colName.index(reportEnd.strftime('%B-%Y')),colName.index(reportEnd.strftime('%B-%Y'))-12,-1): 
        df_SRR['DAYS IN LAST 12'] += df_SRR[colName[i]]  
    
    #CALULA LOS DIAS EN LOS ULTIMOS 5 MESES, SE USA CUANDO ES 6  MONTH ROLLING Y POR EL MAXA NO LLEGA
    df_SRR['DAYS IN LAST 5'] = 0
    for i in range(colName.index(reportEnd.strftime('%B-%Y')),colName.index(reportEnd.strftime('%B-%Y'))-5,-1): 
        df_SRR['DAYS IN LAST 5'] += df_SRR[colName[i]]    
    
    #CALULA LOS DIAS EN LOS ULTIMOS 6 MESES, SE USA CUANDO ES CY Y POR EL MAXA NO LLEGA
    df_SRR['DAYS IN LAST 6'] = 0
    for i in range(colName.index(reportEnd.strftime('%B-%Y')),colName.index(reportEnd.strftime('%B-%Y'))-6,-1): 
        df_SRR['DAYS IN LAST 6'] += df_SRR[colName[i]]   
    
    df_SRR['Total Days'] = df_SRR['12 TY']
    df_SRR['Total Days_Final'] = df_SRR['12 TY']   
    
    df_SRR.loc[df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling','Total Days'] = df_SRR['MAXA']

    df_SRR.loc[df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling','Total Days_Final'] = df_SRR['MAXA']

    
    
    df_SRR.loc[(df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling') &
                (df_SRR['MAXA']<df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered']) &
                (df_SRR['Destination Country']!='Austria') &
                (df_SRR['Destination Country']!='Czech Republic') & (df_SRR['Destination Country']!='Slovakia'),'Total Days'] = df_SRR['DAYS IN LAST 12']
    
    df_SRR.loc[(df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling') &
                (df_SRR['MAXA']<df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered']) &
                (df_SRR['Destination Country']!='Austria') &
                (df_SRR['Destination Country']!='Czech Republic') & (df_SRR['Destination Country']!='Slovakia'),'Total Days_Final'] = df_SRR['DAYS IN LAST 11']
    
    
    df_SRR.loc[(df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling') &
                (df_SRR['MAXA']<df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered']) &
                ((df_SRR['Destination Country']=='Austria') |
                (df_SRR['Destination Country']=='Czech Republic') | (df_SRR['Destination Country']=='Slovakia')),'Total Days'] = df_SRR['DAYS IN LAST 6']
                
                
    
    df_SRR.loc[(df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling') &
                (df_SRR['MAXA']<df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered']) &
                ((df_SRR['Destination Country']=='Austria') |
                (df_SRR['Destination Country']=='Czech Republic') | (df_SRR['Destination Country']=='Slovakia')),'Total Days_Final'] = df_SRR['DAYS IN LAST 5']
                
                
                
                
                
    df_SRR.drop(columns=['12 TY','DAYS IN LAST 11','DAYS IN LAST 5','DAYS IN LAST 12','DAYS IN LAST 6'],inplace=True)
    
    ''' EXCEPTIONS EN EL TOTAL DAYS
    1. UK ROLLING: SUMA 2 AÑOS
    2. SOUTH AFRICA Y NAMIBIA: CY * REGLA PARA LOS NON CY QUE TENGAN REGLA DE CY => HAY QUE SUMAR LA PORCION DE DIAS DENTRO DEL TY Y DEL CY
    '''
    ##### UK DAYS
    df_SRR['UK DAYS'] = 0
    for i in range(24):
        df_SRR['UK DAYS'] += df_SRR[colName[i]]
    df_SRR.loc[(df_SRR['Destination Country']=='United Kingdom')  &
                (df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling'),'Total Days'] = df_SRR['UK DAYS']

    df_SRR.loc[(df_SRR['Destination Country']=='United Kingdom')  &
                (df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='Rolling'),'Total Days_Final'] = df_SRR['UK DAYS']

    df_SRR.drop(columns=['UK DAYS'],inplace=True)
    
    
    ### CY THRESHOLD FOR FY COUNTRIES
    print('Calculating CY Threshold for FY Countries ' + datetime.now().time().strftime('%H:%M:%S'))
    if tyStart.month != 1:
        df_SRR['CY THRESHOLD'] = 0
        posDecember = colName.index('December-' + str(tyStart.year))
        posPeriod = colName.index(reportEnd.strftime('%B-%Y'))
        if posDecember > posPeriod:
            for i in range(12,posDecember+1):
                df_SRR['CY THRESHOLD'] += df_SRR[colName[i]]
        else:
            for i in range(posDecember+1,24):
                df_SRR['CY THRESHOLD'] += df_SRR[colName[i]]
        
        df_SRR.loc[df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='CY','Total Days'] = df_SRR['CY THRESHOLD']
        df_SRR.loc[df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)']=='CY','Total Days_Final'] = df_SRR['CY THRESHOLD']

        df_SRR.drop(columns=['CY THRESHOLD'],inplace=True)
    
    
    ################## END OF EXCEPTIONS
    
    df_SRR['Tax Threshold'] = df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered'].astype(int).astype(str) + ' ' + df_SRR['Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)'] + ' ' + df_SRR['Day Type']
    
    df_SRR['Days to Reach Taxability'] = df_SRR['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered'] - df_SRR['Total Days_Final']
    
    df_SRR.drop(columns=['Income Tax Threshold: \nFirst Day at Which Income Tax Is Triggered','Income Tax Threshold:  \nPeriod \n(CY, Rolling, TY, \nNo income tax)','Total Days_Final'],inplace=True)  
    #MODIFICADO POR NAHUEL.A.RIOS 4/1/20
    return df_SRR,df_LBD
    
def Status(df_LBD,df_SRR,reportStart,reportEnd):
    print('Analyzing Status columns ' + datetime.now().time().strftime('%H:%M:%S'))
    df_PP = DataFrame.pivot_table(df_LBD,values='DateVal',index=['People Key'],columns=['Period'],aggfunc=lambda x: len(x.unique())).fillna(0)
    
    #DA ORDEN A LAS COLUMNAS
    colDate = reportStart.replace(day=1)
    colName = []
    for _ in range(24):
        colName.append('Status ' + colDate.strftime('%B-%Y'))
        if colDate.strftime('%B-%Y') not in df_PP:
            df_PP['Status ' + colDate.strftime('%B-%Y')] = 'Not Submitted'
        else:
            df_PP.loc[df_PP[colDate.strftime('%B-%Y')]>0,colDate.strftime('%B-%Y')] = 'Submitted'
            df_PP.loc[df_PP[colDate.strftime('%B-%Y')]==0,colDate.strftime('%B-%Y')] = 'Not Submitted'
            df_PP.rename(columns={colDate.strftime('%B-%Y'):'Status ' + colDate.strftime('%B-%Y')},inplace=True)
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
    df_PP = df_PP[colName]
    df_SRR = df_SRR.merge(df_PP,how='left',left_on=['People Key'],right_index=True)
    df_SRR[colName]=df_SRR[colName].fillna('Not Submitted')
    
    df_SRR['Missing MyTE Period Within Assignnment'] = 0
    colDate = reportStart.replace(day=1)
    for _ in range(24):
        col = ('Status ' + colDate.strftime('%B-%Y'))
        if colDate < reportEnd:
            df_SRR.loc[df_SRR[col]=='Not Submitted','Missing MyTE Period Within Assignnment'] += 1
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
    df_SRR.loc[df_SRR['Missing MyTE Period Within Assignnment']==0,'Missing MyTE Period Within Assignnment'] = 'All periods within assignment submitted'
    return df_SRR

def Control(df_SRR,df_LBD,reportStart,reportEnd):
    print('Calculating Control columns ' + datetime.now().time().strftime('%H:%M:%S'))
    #REMUEVE LOS HOME = HOST (TRANSFERS)
    df_Location = df_LBD[(df_LBD['HomeCountry']!=df_LBD['Location1']) & df_LBD['Location1'].isin(df_SRR['Destination Country'].drop_duplicates())]
        
    #PIVOT CON WD Y PIVOT PP
    df_WD = DataFrame.pivot_table(df_Location[(df_Location['TypeOfDay']=='Work Day') & (df_Location['DOA']=='DAYS OUT')],
                                              values='DateVal',index=['People Key','Location1'],columns=['Period'],aggfunc=lambda x: len(x.unique())).fillna(0)
    df_PP = DataFrame.pivot_table(df_Location[df_Location['DOA']=='DAYS OUT'],
                                  values='DateVal' ,index=['People Key','Location1'],columns=['Period'],aggfunc=lambda x: len(x.unique())).fillna(0)
    
    #DA ORDEN A LAS COLUMNAS
    colDate = reportStart.replace(day=1)
    colName = []
    for i in range(24):
        colName.append('Control ' + colDate.strftime('%B-%Y'))
        if colDate.strftime('%B-%Y') not in df_PP:
            df_PP['Control ' + colDate.strftime('%B-%Y')] = None        
        else:
            df_PP.loc[df_PP[colDate.strftime('%B-%Y')]>0,colDate.strftime('%B-%Y')] = 'Days out of assignment: ' + df_PP[colDate.strftime('%B-%Y')].astype(str)
            df_PP.loc[df_PP[colDate.strftime('%B-%Y')]==0,colDate.strftime('%B-%Y')] = None
            df_PP.rename(columns={colDate.strftime('%B-%Y'):'Control ' + colDate.strftime('%B-%Y')},inplace=True)
        if colDate.strftime('%B-%Y') not in df_WD:
            df_WD['Control ' + colDate.strftime('%B-%Y')] = None        
        else:
            df_WD.loc[df_WD[colDate.strftime('%B-%Y')]>0,colDate.strftime('%B-%Y')] = 'Days out of assignment: ' + df_WD[colDate.strftime('%B-%Y')].astype(str)
            df_WD.loc[df_WD[colDate.strftime('%B-%Y')]==0,colDate.strftime('%B-%Y')] = None
            df_WD.rename(columns={colDate.strftime('%B-%Y'):'Control ' + colDate.strftime('%B-%Y')},inplace=True)
        #PONE N/A SI ES POSTERIOR AL PERIODO DE ANALYSIS
        if colDate > reportEnd:
            df_PP['Control ' + colDate.strftime('%B-%Y')] = 'N/A'
            df_WD['Control ' + colDate.strftime('%B-%Y')] = 'N/A'
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
        
    df_PP = df_PP[colName]
    df_WD = df_WD[colName]
    #INCLUYE TRY POR SI NO HAY CASOS CON DAYS OUT
    try:
        df_1 = df_SRR[df_SRR['Day Type']=='PP'].merge(df_PP,how='left',left_on=['People Key','Destination Country'],right_index=True)
    except:
        df_1 = df_SRR[df_SRR['Day Type']=='PP']
    try:
        df_2 = df_SRR[df_SRR['Day Type']=='WD'].merge(df_WD,how='left',left_on=['People Key','Destination Country'],right_index=True)
    except:
        df_2 = df_SRR[df_SRR['Day Type']=='WD']
    df_SRR = df_1.append(df_2)
    #SI NO HAY CASOS CON DAYS OUT INCLUYE LAS COLUMNAS VACIAS
    if len(set(colName)-set(df_SRR))>0:
        for c in colName:
            df_SRR[c] = None
    
    #HACE EL CONTROL PARA EL RESTO DE LOS CASOS
    df_SRR['0 Days Within Assignment (number of period with 0 days)'] = 0
    for i in range(24):
        dt_1 = reportStart.replace(day=1) + DateOffset(months=i)
        if dt_1 <= reportEnd:
            dt_2 = reportStart.replace(day=1) + DateOffset(months=i+1) - timedelta(days=1)
            colDias = dt_1.strftime('%B-%Y')
            colControl = 'Control ' + dt_1.strftime('%B-%Y')
            df_SRR.loc[(df_SRR[colControl].isnull()) &
                        (df_SRR['Travel Start Date']<=dt_2) & (df_SRR['Travel End Date']>=dt_1) &
                        (df_SRR[colDias]==0),colControl] = '0 days within assignment'
            df_SRR.loc[(df_SRR[colControl]=='0 days within assignment'),'0 Days Within Assignment (number of period with 0 days)'] +=1
            df_SRR.loc[(df_SRR[colControl].isnull()) & (df_SRR[colDias]>0),colControl] = 'OK'
            df_SRR.loc[(df_SRR[colControl].isnull()),colControl] = 'No assignment'
    df_SRR.loc[df_SRR['0 Days Within Assignment (number of period with 0 days)']==0,'0 Days Within Assignment (number of period with 0 days)'] = 'All periods within assignment with days'
    
    #Days Out of Assignment (total number of days out)
    #PIVOT CON WD Y PIVOT PP
    df_WD = DataFrame.pivot_table(df_Location[(df_Location['TypeOfDay']=='Work Day') & (df_Location['DOA']=='DAYS OUT')],
                                              values='DateVal',index=['People Key','Location1'],aggfunc=lambda x: len(x.unique())).fillna(0)
    df_PP = DataFrame.pivot_table(df_Location[df_Location['DOA']=='DAYS OUT'],
                                  values='DateVal' ,index=['People Key','Location1'],aggfunc=lambda x: len(x.unique())).fillna(0)
    df_PP.rename(columns={'DateVal':'Days Out of Assignment (total number of days out)'},inplace=True)
    df_WD.rename(columns={'DateVal':'Days Out of Assignment (total number of days out)'},inplace=True)
    try:
        df_1 = df_SRR[df_SRR['Day Type']=='PP'].merge(df_PP,how='left',left_on=['People Key','Destination Country'],right_index=True)
    except:
        df_1 = df_SRR[df_SRR['Day Type']=='PP']
    try:
        df_2 = df_SRR[df_SRR['Day Type']=='WD'].merge(df_WD,how='left',left_on=['People Key','Destination Country'],right_index=True)
    except:
        df_2 = df_SRR[df_SRR['Day Type']=='WD']
    df_SRR = df_1.append(df_2)
    if 'Days Out of Assignment (total number of days out)' in df_SRR:
        df_SRR['Days Out of Assignment (total number of days out)'].fillna('No days out of assignment',inplace=True)
    else:
        df_SRR['Days Out of Assignment (total number of days out)'] = 'No days out of assignment'
    
    
    return df_SRR

def Arrange(df_SRR,Col_SRR,reportStart,reportEnd,tyStart):
    print('Arranging columns ' + datetime.now().time().strftime('%H:%M:%S'))
    for i in range(24):
        Col_SRR.append((reportStart.replace(day=1) + DateOffset(months=i)).strftime('%B-%Y'))
        Col_SRR.append('Status ' + (reportStart.replace(day=1) + DateOffset(months=i)).strftime('%B-%Y'))
        Col_SRR.append('Control ' + (reportStart.replace(day=1) + DateOffset(months=i)).strftime('%B-%Y'))
    
    Col_SRR.append('Total Days')
    Col_SRR.append('Tax Threshold')
    Col_SRR.append('Days to Reach Taxability')
    Col_SRR.append('Missing MyTE Period Within Assignnment')
    Col_SRR.append('Days Out of Assignment (total number of days out)')
    Col_SRR.append('0 Days Within Assignment (number of period with 0 days)')
    Col_SRR.append('Taxability This Month')
    Col_SRR.append('Taxability Next Month')
    Col_SRR.append('MAXA')
    
    #TAXABILITY COLUMNS
    df_SRR['Taxability This Month'] = 'Not Taxable'
    df_SRR.loc[(df_SRR['Taxable End Date']>=tyStart) | (df_SRR['Days to Reach Taxability']<=0),'Taxability This Month'] = 'Taxable'
    df_SRR['Taxability Next Month'] = 'May Not Reach Taxability'
    df_SRR.loc[df_SRR['Taxability This Month']=='Taxable','Taxability Next Month'] = 'Taxable'
    df_SRR.loc[(df_SRR['Taxability Next Month']=='May Not Reach Taxability') & (df_SRR['End Date Confirmed']=='Yes'),'Taxability Next Month'] = 'Not Taxable'
    days_next_month = (reportEnd.replace(day=1) + DateOffset(months=2) - timedelta(days=1)).day
    df_SRR.loc[(df_SRR['Taxability Next Month']=='May Not Reach Taxability') & (df_SRR['End Date Confirmed']=='No') &
                (df_SRR['Days to Reach Taxability']<=days_next_month),'Taxability Next Month'] = 'May Reach Taxability'
    
    df_SRR = df_SRR[Col_SRR]
    
    colDate = reportStart.replace(day=1)
    for i in range(24):
        colDate = reportStart.replace(day=1) + DateOffset(months=i)
        if colDate > reportEnd:
            df_SRR[colDate.strftime('%B-%Y')] = None
            df_SRR['Status ' + colDate.strftime('%B-%Y')] = None
            df_SRR['Control ' + colDate.strftime('%B-%Y')] = None
        
    return df_SRR

def Percents(df_SRR,df_LBD,tyStart):
    print('Calculating percents ' + datetime.now().time().strftime('%H:%M:%S'))
    #Host PP Days in Current TY
    print('Host PP Days in Current TY ' + datetime.now().time().strftime('%H:%M:%S'))
    pivot = DataFrame.pivot_table(df_LBD[(df_LBD['HomeCountry']!=df_LBD['Location1']) & (df_LBD['DateVal']>=tyStart)],
                                  values='DateVal',index=['People Key','Location1'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    if len(pivot) != 0:
        pivot.rename(columns={'DateVal':'Host PP Days in Current TY'},inplace=True)
        df_SRR = df_SRR.merge(pivot,how='left',left_on=['People Key','Destination Country'],right_index=True)
    if 'Host PP Days in Current TY' not in df_SRR: df_SRR['Host PP Days in Current TY'] = 0
    
    #Total PP Days in Current TY
    print('Total PP Days in Current TY ' + datetime.now().time().strftime('%H:%M:%S'))
    pivot = DataFrame.pivot_table(df_LBD[(df_LBD['DateVal']>=tyStart)],
                                  values='DateVal',index=['People Key'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    if len(pivot) != 0:
        pivot.rename(columns={'DateVal':'Total PP Days in Current TY'},inplace=True)
        df_SRR = df_SRR.merge(pivot,how='left',left_on=['People Key'],right_index=True)
    if 'Total PP Days in Current TY' not in df_SRR: df_SRR['Total PP Days in Current TY'] = 0
    
    #Percent PP Days in Current TY
    print('Percent PP Days in Current TY ' + datetime.now().time().strftime('%H:%M:%S'))
    df_SRR['Percent PP Days in Current TY'] = df_SRR['Host PP Days in Current TY'] / df_SRR['Total PP Days in Current TY']
    df_SRR['Percent PP Days in Current TY'].fillna(0)
    
    #Host WD Days in Current TY
    print('Host WD Days in Current TY ' + datetime.now().time().strftime('%H:%M:%S'))
    pivot = DataFrame.pivot_table(df_LBD[(df_LBD['HomeCountry']!=df_LBD['Location1']) & (df_LBD['DateVal']>=tyStart) & (df_LBD['TypeOfDay']=='Work Day')],
                                  values='DateVal',index=['People Key','Location1'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    if len(pivot) != 0:
        pivot.rename(columns={'DateVal':'Host WD Days in Current TY'},inplace=True)
        df_SRR = df_SRR.merge(pivot,how='left',left_on=['People Key','Destination Country'],right_index=True)
    if 'Host WD Days in Current TY' not in df_SRR: df_SRR['Host WD Days in Current TY'] = 0
    
    #Total WD Days in Current TY
    print('Total WD Days in Current TY ' + datetime.now().time().strftime('%H:%M:%S'))
    pivot = DataFrame.pivot_table(df_LBD[(df_LBD['DateVal']>=tyStart) & (df_LBD['TypeOfDay']=='Work Day')],
                                  values='DateVal',index=['People Key'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    if len(pivot) != 0:
        pivot.rename(columns={'DateVal':'Total WD Days in Current TY'},inplace=True)
        df_SRR = df_SRR.merge(pivot,how='left',left_on=['People Key'],right_index=True)
    if 'Total WD Days in Current TY' not in df_SRR: df_SRR['Total WD Days in Current TY'] = 0
    
    #Percent WD Days in Current TY
    print('Percent WD Days in Current TY ' + datetime.now().time().strftime('%H:%M:%S'))
    df_SRR['Percent WD Days in Current TY'] = df_SRR['Host WD Days in Current TY'] / df_SRR['Total WD Days in Current TY']
    df_SRR['Percent WD Days in Current TY'].fillna(0)
    
    #Host PP Days in Current Assignment
    print('Host PP Days in Current Assignment ' + datetime.now().time().strftime('%H:%M:%S'))
    df = df_SRR[['People Key','Destination Country','Travel Plan Number','Travel Start Date','Travel End Date']]
    df = df.merge(df_LBD[['People Key','Location1','DateVal']][df_LBD['HomeCountry']!=df_LBD['Location1']],
                  how='inner',left_on=['People Key','Destination Country'],right_on=['People Key','Location1']).drop_duplicates()
    df['WITHIN'] = False
    df.loc[(df['DateVal']>=df['Travel Start Date']) & (df['DateVal']<=df['Travel End Date']),'WITHIN'] = True
    pivot = DataFrame.pivot_table(df[df['WITHIN']==True],
                                  values='DateVal',index=['Travel Plan Number'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    pivot.rename(columns={'DateVal':'Host PP Days in Current Assignment'},inplace=True)
    df_SRR = df_SRR.merge(pivot,how='left',left_on=['Travel Plan Number'],right_index=True)
    if 'Host PP Days in Current Assignment' not in df_SRR: df_SRR['Host PP Days in Current Assignment'] = 0
    
    #Total PP Days in Current Assignment
    print('Total PP Days in Current Assigment ' + datetime.now().time().strftime('%H:%M:%S'))
    df = df_SRR[['People Key','Destination Country','Travel Plan Number','Travel Start Date','Travel End Date']]
    df = df.merge(df_LBD[['People Key','DateVal']],
                  how='inner',left_on=['People Key'],right_on=['People Key']).drop_duplicates()
    df['WITHIN'] = False
    df.loc[(df['DateVal']>=df['Travel Start Date']) & (df['DateVal']<=df['Travel End Date']),'WITHIN'] = True
    pivot = DataFrame.pivot_table(df[df['WITHIN']==True],
                                  values='DateVal',index=['Travel Plan Number'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    pivot.rename(columns={'DateVal':'Total PP Days in Current Assignment'},inplace=True)
    df_SRR = df_SRR.merge(pivot,how='left',left_on=['Travel Plan Number'],right_index=True)
    if 'Total PP Days in Current Assignment' not in df_SRR: df_SRR['Total PP Days in Current Assignment'] = 0
    
    #Percent PP Days in Current Assignment
    print('Percent PP Days in Current Assignment ' + datetime.now().time().strftime('%H:%M:%S'))
    df_SRR['Percent PP Days in Current Assignment'] = df_SRR['Host PP Days in Current Assignment'] / df_SRR['Total PP Days in Current Assignment']
    df_SRR['Percent PP Days in Current Assignment'].fillna(0)
    
    #Host WD Days in Current Assignment
    print('Host WD Days in Current Assignment ' + datetime.now().time().strftime('%H:%M:%S'))
    df = df_SRR[['People Key','Destination Country','Travel Plan Number','Travel Start Date','Travel End Date']]
    df = df.merge(df_LBD[['People Key','Location1','DateVal']][(df_LBD['HomeCountry']!=df_LBD['Location1']) & (df_LBD['TypeOfDay']=='Work Day')],
                  how='inner',left_on=['People Key','Destination Country'],right_on=['People Key','Location1']).drop_duplicates()
    df['WITHIN'] = False
    df.loc[(df['DateVal']>=df['Travel Start Date']) & (df['DateVal']<=df['Travel End Date']),'WITHIN'] = True
    pivot = DataFrame.pivot_table(df[df['WITHIN']==True],
                                  values='DateVal',index=['Travel Plan Number'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    pivot.rename(columns={'DateVal':'Host WD Days in Current Assignment'},inplace=True)
    df_SRR = df_SRR.merge(pivot,how='left',left_on=['Travel Plan Number'],right_index=True)
    if 'Host WD Days in Current Assignment' not in df_SRR: df_SRR['Host WD Days in Current Assignment'] = 0
    
    #Total WD Days in Current Assignment
    print('Total WD Days in Current Assignment ' + datetime.now().time().strftime('%H:%M:%S'))
    df = df_SRR[['People Key','Destination Country','Travel Plan Number','Travel Start Date','Travel End Date']]
    df = df.merge(df_LBD[['People Key','DateVal']][df_LBD['TypeOfDay']=='Work Day'],
                  how='inner',left_on=['People Key'],right_on=['People Key']).drop_duplicates()
    df['WITHIN'] = False
    df.loc[(df['DateVal']>=df['Travel Start Date']) & (df['DateVal']<=df['Travel End Date']),'WITHIN'] = True
    pivot = DataFrame.pivot_table(df[df['WITHIN']==True],
                                  values='DateVal',index=['Travel Plan Number'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    pivot.rename(columns={'DateVal':'Total WD Days in Current Assignment'},inplace=True)
    df_SRR = df_SRR.merge(pivot,how='left',left_on=['Travel Plan Number'],right_index=True)
    if 'Total WD Days in Current Assignment' not in df_SRR: df_SRR['Total WD Days in Current Assignment'] = 0
    
    #Percent WD Days in Current Assignment
    print('Percent WD Days in Current Assignment ' + datetime.now().time().strftime('%H:%M:%S'))
    df_SRR['Percent WD Days in Current Assignment'] = df_SRR['Host WD Days in Current Assignment'] / df_SRR['Total WD Days in Current Assignment']
    df_SRR['Percent WD Days in Current Assignment'].fillna(0)
    
    return df_SRR

def DOA_Analysis(df_SRR,df_LBD,tyStart):
    print('Final DOA analysis ' + datetime.now().time().strftime('%H:%M:%S'))
    
    #Days Out Before
    df = df_SRR[['People Key','Destination Country','Travel Plan Number','Travel Start Date','Travel End Date']]
    df = df.merge(df_LBD[['People Key','Location1','DateVal']][(df_LBD['HomeCountry']!=df_LBD['Location1']) & (df_LBD['DOA']=='DAYS OUT')],
                  how='inner',left_on=['People Key','Destination Country'],right_on=['People Key','Location1']).drop_duplicates()
    df['DOA Analysis'] = None
    df.loc[(df['DOA Analysis'].isnull()) &
            (df['DateVal']<tyStart),'DOA Analysis'] = 'Days Out Before ' + tyStart.strftime('%d-%b-%Y')
    df.loc[(df['DOA Analysis'].isnull()) &
            (df['DateVal']>=df['Travel Start Date'] - timedelta(days=15)) & (df['DateVal']<df['Travel Start Date']),'DOA Analysis'] = 'Days Out Within 15 to Start Date'
    df.loc[(df['DOA Analysis'].isnull()) &
            (df['DateVal']>=df['Travel Start Date'] - timedelta(days=30)) & (df['DateVal']<df['Travel Start Date']),'DOA Analysis'] = 'Days Out Within 30 to Start Date'
    df.loc[(df['DOA Analysis'].isnull()) &
            (df['DateVal']<df['Travel Start Date']),'DOA Analysis'] = 'Days Out Over 30 to Start Date'
    df.loc[(df['DOA Analysis'].isnull()) &
            (df['DateVal']<=df['Travel End Date'] + timedelta(days=15)) & (df['DateVal']>df['Travel End Date']),'DOA Analysis'] = 'Days Out Within 15 to End Date'
    df.loc[(df['DOA Analysis'].isnull()) &
            (df['DateVal']<=df['Travel End Date'] + timedelta(days=30)) & (df['DateVal']>df['Travel End Date']),'DOA Analysis'] = 'Days Out Within 30 to End Date'
    df.loc[(df['DOA Analysis'].isnull()) &
            (df['DateVal']>df['Travel End Date']),'DOA Analysis'] = 'Days Out Over 30 to End Date'        
    pivot = DataFrame.pivot_table(df, columns=['DOA Analysis'], 
                                  values='DateVal',index=['Travel Plan Number'],
                                  aggfunc=lambda x: len(x.unique())).fillna(0)
    #ORDENA LAS COLUMNAS
    if 'Days Out Before ' + tyStart.strftime('%d-%b-%Y') not in pivot: pivot['Days Out Before ' + tyStart.strftime('%d-%b-%Y')] = 0
    if 'Days Out Within 15 to Start Date' not in pivot: pivot['Days Out Within 15 to Start Date'] = 0
    if 'Days Out Within 15 to End Date' not in pivot: pivot['Days Out Within 15 to End Date'] = 0
    if 'Days Out Within 30 to Start Date' not in pivot: pivot['Days Out Within 30 to Start Date'] = 0
    if 'Days Out Within 30 to End Date' not in pivot: pivot['Days Out Within 30 to End Date'] = 0
    if 'Days Out Over 30 to Start Date' not in pivot: pivot['Days Out Over 30 to Start Date'] = 0
    if 'Days Out Over 30 to End Date' not in pivot: pivot['Days Out Over 30 to End Date'] = 0
    pivot = pivot[['Days Out Before ' + tyStart.strftime('%d-%b-%Y'),
                  'Days Out Within 15 to Start Date','Days Out Within 15 to End Date',
                  'Days Out Within 30 to Start Date', 'Days Out Within 30 to End Date',
                  'Days Out Over 30 to Start Date','Days Out Over 30 to End Date']]
                  
    df_SRR = df_SRR.merge(pivot,how='left',left_on=['Travel Plan Number'],right_index=True)
    
    return df_SRR

def Final_Report(df,list_ARE,reportStart,reportEnd,tyStart,srrType,deskPath,srrName,Countries,GU,midSRR,manualInput):
    def Give_Format(ws):
        my_red = PatternFill(patternType='solid', fgColor=Color(rgb='00FF0000'))
        my_silver = PatternFill(patternType='solid', fgColor=Color(rgb='00C0C0C0'))
        my_orange = PatternFill(patternType='solid', fgColor=Color(rgb='ff9f32'))
        my_darkgreen = PatternFill(patternType='solid', fgColor=Color(rgb='177700'))
        my_green = PatternFill(patternType='solid', fgColor=Color(rgb='59b769'))
        my_white = PatternFill(patternType='solid', fgColor=Color(rgb='00FFFFFF'))
        my_yellow = PatternFill(patternType='solid', fgColor=Color(rgb='fffb91'))
        my_darkblue = PatternFill(patternType='solid', fgColor=Color(rgb='001e4f'))
        
        #FORMATO A ENCABEZADOS
        for i in range(len(colName)):
            cell = ws.cell(row=1,column=i+1)
            cell.font = Font(color=colors.WHITE)
            cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text = True)
        for i in range(colName.index('SRR Comments')):
            ws.cell(row=1,column=i+1).fill = my_darkblue
        ws.cell(row=1,column=colName.index('SRR Comments')+1).fill = my_orange
        for i in range(colName.index(reportStart.strftime('%B-%Y')),colName.index(reportStart.strftime('%B-%Y'))+24*3):
            ws.cell(row=1,column=i+1).fill = my_darkgreen
        for i in range(colName.index('Total Days'),colName.index('Days to Reach Taxability')+1):
            ws.cell(row=1,column=i+1).fill = my_silver
        for i in range(colName.index('Missing MyTE Period Within Assignnment'),colName.index('MAXA')+1):
            ws.cell(row=1,column=i+1).fill = my_darkgreen
        for i in range(colName.index('Host PP Days in Current TY'),colName.index('Percent WD Days in Current Assignment')+1):
            ws.cell(row=1,column=i+1).fill = my_orange
        for i in range(colName.index('Percent WD Days in Current Assignment')+1,colName.index('Days Out Over 30 to End Date')+1):
            ws.cell(row=1,column=i+1).fill = my_yellow
            ws.cell(row=1,column=i+1).font = Font(color=colors.BLACK)
            
        min_col = colName.index('Control ' + reportStart.strftime('%B-%Y')) + 1
        max_col = min_col
        last_col = colName.index('Control ' + reportEnd.strftime('%B-%Y')) + 1
        
        min_row=2; max_row = ws.max_row        
        #COMPLETA COLOR EN BASE A LAS COLUMNAS DE CONTROL
        while min_col <= last_col:
            for col in ws.iter_cols(min_col,min_col,min_row,max_row):
                for cell in col: 
                    if cell.value == "OK":
                        cell.fill = my_green
                        celda = cell.offset(0,-1)
                        celda.fill = my_green
                        celda = cell.offset(0,-2)
                        celda.fill = my_green
                    elif cell.value == "No assignment":
                        cell.fill = my_silver
                        celda = cell.offset(0,-1)
                        celda.fill = my_silver
                        celda = cell.offset(0,-2)
                        celda.fill = my_silver
                    elif cell.value == "0 days within assignment":
                        cell.fill = my_orange
                        celda = cell.offset(0,-1)
                        celda.fill = my_orange
                        celda = cell.offset(0,-2)
                        celda.fill = my_orange
                    elif cell.value is None:
                        cell.fill = my_white
                        celda = cell.offset(0,-1)
                        celda.fill = my_white
                        celda = cell.offset(0,-2)
                        celda.fill = my_white          
                    else:
                        cell.fill = my_red
                        celda = cell.offset(0,-1)
                        celda.fill = my_red
                        celda = cell.offset(0,-2)
                        celda.fill = my_red   
                    
            min_col = min_col + 3
            # max_col = max_col + 3
            min_row=2
        
        #PINTA LAS 3 DE STATUS
        min_col = colName.index('Missing MyTE Period Within Assignnment') + 1
        max_col = min_col
        last_col = colName.index('0 Days Within Assignment (number of period with 0 days)') + 1
        
        while max_col <= last_col:
            for col in ws.iter_cols(min_col,max_col,min_row,max_row):
                for cell in col: 
                    if (cell.value == 'All periods within assignment submitted') | (cell.value == 'No days out of assignment') | (cell.value == 'All periods within assignment with days') :
                        cell.fill = my_green      
                    else:
                        cell.fill = my_red         
            min_col = min_col + 1
            max_col = max_col + 1
            min_row=2
            
        min_col = 1
        min_col = colName.index('Current Employee ID')
        max_col = colName.index('Company Code') + 1
        min_row= 2
        max_row = ws.max_row  
                
        r = 2
        for rows in ws.iter_rows(min_row, max_row, min_col, max_col):
            company_code = ws.cell(row = r , column = colName.index('Company Code') + 1).value
            r = r +1
            if company_code in list_ARE:
               for cell in rows:
                   cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb='ff9f32'))
        return ws
    
    colName = list(df.columns)
    
    df.replace({pd.NaT: ''}, inplace=True)
    
    wb = Workbook()
    ws = wb.active
    normal = wb._named_styles['Normal']
    normal.font.size = 10
    for dataframe in dataframe_to_rows(df, index=False, header=True):
            ws.append(dataframe)
    
    print('Formating SRR ' + datetime.now().time().strftime('%H:%M:%S'))
    ws = Give_Format(ws)
    ws.insert_rows(1,3)
    ws.cell(row=2,column=2).value = srrName
    ws.cell(row=2,column=2).font = Font(bold=True,size=16)
    
    print('Exporting SRR ' + datetime.now().time().strftime('%H:%M:%S'))
    ws.title = 'SRR'
    
    ws_inputs = wb.create_sheet()
    list_fields = ['Field','GU of Analysis','Country of Analysis','Tax Year Start Date','Month of Analysis','SRR Type','Mid SRR','Manual Input']
    if midSRR > 0:
        midSRR = True
    else:
        midSRR = False
    if manualInput > 0:
        manualInput = True
    else:
        manualInput = False
        
        

    list_inputs = ['Input',GU,','.join(Countries),tyStart,reportEnd,srrType,midSRR,manualInput]
    for i in range(8):
        ws_inputs.cell(row = i + 1, column = 1).value = list_fields[i]
        ws_inputs.cell(row = i + 1, column = 2).value = list_inputs[i]
    ws_inputs.title = 'Inputs'
    
    img = drawing.image.Image(deskPath + r'\Supporting Doc\Input_Parameters.png')
    img.anchor = 'D1'
    img.width = 800
    img.height = 600
    ws_inputs.add_image(img)

    wb.save(deskPath + '\\' + srrName + '.xlsx')
    
    print('Preparing IC10 ' + datetime.now().time().strftime('%H:%M:%S'))
    df['Comments']= None
    if srrType == 'Compliance':
        df = df[(df['Tax Compliance (Destination)']=='Yes') | (df['Days to Reach Taxability']<=0)]
        df['Compliance Status for TY of Analysis'] = None
        df.loc[df['Travel End Date']<tyStart,['Comments','Compliance Status for TY of Analysis']] = ['Assignment for Previous Tax Year','Not Compliance']
    else:
        df['Taxable Status for TY of Analysis'] = None
        try:
            df_1 = df[df['Taxable In Host']=='Yes']
            df_1 = df_1[df_1['Taxable End Date']>=tyStart]
            df_1.loc[df_1['Days to Reach Taxability']<=0,['Comments','Taxable Status for TY of Analysis']] = ['Already Updated as Taxable','Taxable']
            df_2 = df[(df['Days to Reach Taxability']<=0)]
            df = df_1.append(df_2)
            df.drop_duplicates(['Travel Plan Number'],inplace=True)
        except:
            pass
        df.loc[df['Travel End Date']<tyStart,['Comments','Taxable Status for TY of Analysis']] = ['Assignment for Previous Tax Year','Not Taxable']
    
    df.replace({pd.NaT: ''}, inplace=True)
    
    ws_IC = wb.create_sheet()
    for dataframe in dataframe_to_rows(df, index=False, header=True):
        ws_IC.append(dataframe)
    print('Formating IC10 ' + datetime.now().time().strftime('%H:%M:%S'))
    ws_IC = Give_Format(ws_IC)
    
    for i in range(colName.index('Days Out Over 30 to End Date'),colName.index('Missing MyTE Period Within Assignnment')-1,-1):
        ws_IC.delete_cols(i+1)
    my_blue = PatternFill(patternType='solid', fgColor=Color(rgb='aabeff')) 
    ws_IC.cell(row=1,column=colName.index('Missing MyTE Period Within Assignnment')+1).fill = my_blue 
    ws_IC.cell(row=1,column=colName.index('Missing MyTE Period Within Assignnment')+2).fill = my_blue
    ws_IC.cell(row=1,column=colName.index('Missing MyTE Period Within Assignnment')+1).alignment = Alignment(horizontal='center',vertical='center',wrap_text = True) 
    ws_IC.cell(row=1,column=colName.index('Missing MyTE Period Within Assignnment')+2).alignment = Alignment(horizontal='center',vertical='center',wrap_text = True)
    for i in range(colName.index('Total Days')-1,colName.index('SRR Comments')-1,-1):
        if 'SRR Comments' in colName[i] or 'Control' in colName[i] or 'Status' in colName[i]:
            ws_IC.delete_cols(i+1)
    
    max_row=ws_IC.max_row+2
    ws_IC.cell(row=max_row,column=1).value = 'Actioned By:'
    ws_IC.cell(row=max_row+1,column=1).value = 'Date:'
    ws_IC.cell(row=max_row+3,column=1).value = 'Reviewed By:'
    ws_IC.cell(row=max_row+4,column=1).value = 'Date:'
    
    ws_IC.insert_rows(1,3)
    ws_IC.cell(row=2,column=2).value = srrName.replace('SRR','IC10')
    ws_IC.cell(row=2,column=2).font = Font(bold=True,size=16)
    
    
    print('Exporting IC10 ' + datetime.now().time().strftime('%H:%M:%S'))
    ws_IC.title = 'IC10'
    wb.save(deskPath + '\\' + srrName.replace('SRR','IC10') + '.xlsx')
    
    return

class srrGUI(Tk):
    def __init__(self,BD):
        Tk.__init__(self)
                
        #FRAME PRINCIPAL
        self.main = Frame(self,width=100)
        self.main.pack(side='top')
        
        label = Label(self.main,
                      text='SRR Tool ' + BD.version,
                      font=("Prusia", 14,"italic","bold","underline"))
        label.grid(row=0,column=0,columnspan=3)
        
        #DROPDOWN DE GUs
        s_GU = BD.df_CountryList['Geographic Unit Description'].drop_duplicates().sort_values().reset_index(drop=True)
        lablel = Label(self.main,
                     text='Select GU of Analysis:')
        lablel.grid(row=1,column=0,
                  padx=10,pady=10,
                  sticky=W+E)
        #DROP DOWN
        self.varGU = StringVar(self.main)
        self.varGU.set(s_GU[0])
        self.optGU = OptionMenu(self.main,self.varGU,*s_GU)
        self.optGU.config(width=40)
        self.optGU.grid(row=1,column=1,columnspan=2,
                 padx=10,pady=10,
                 sticky=W+E)
        self.varGU.trace('w', self.change_GU)
                
        #DROPDOWN TY START
        s_TY = BD.df_CountryList[(BD.df_CountryList['Geographic Unit Description']==self.varGU.get())]['Tax Year Start Date'].drop_duplicates().sort_values().reset_index(drop=True)
        lablel = Label(self.main,
                     text='Select Tax Year Start:')
        lablel.grid(row=2,column=0,
                  padx=10,pady=10,
                  sticky=W+E)
        #DROP DOWN
        self.varTY = StringVar(self.main)
        self.varTY.set(s_TY[0])
        self.opTY = OptionMenu(self.main,self.varTY,*s_TY)
        self.opTY.config(width=15)
        self.opTY.grid(row=2,column=1,
                 padx=10,pady=10,
                 sticky=W+E) 
        self.varTY.trace('w',self.change_TY)
        
        #DROPDOWN AÑO
        iYear = date.today().replace(day=1)
        iYear = iYear - timedelta(days=1)
        iYear = iYear.year
        i = -3
        s_Year = []
        while iYear+i <= date.today().year:
            s_Year.append(iYear+i)
            i+=1
        self.varYear = StringVar(self.main)
        self.varYear.set(iYear)
        self.opYear = OptionMenu(self.main,self.varYear,*s_Year)
        self.opYear.config(width=15)
        self.opYear.grid(row=2,column=2,
                 padx=10,pady=10,
                 sticky=W+E)     
        
        self.main.grid_rowconfigure(3,minsize=50)
        
        actionButton = Button(self.main,
                             text='Run SRR/IC10',
                             command=self.actionButtonPress)
        actionButton.grid(row=7,column=0,columnspan=3,sticky=N+S+W+E)
        
        #LIST BOX PAISES
        self.s_Paises = BD.df_CountryList[(BD.df_CountryList['Geographic Unit Description']==self.varGU.get()) &
                              (BD.df_CountryList['Tax Year Start Date']==self.varTY.get())]['Country Name'].drop_duplicates().sort_values().reset_index(drop=True)
        scrollbar = Scrollbar(self.main, orient='vertical')
        self.paises = Listbox(self.main,selectmode='extended', yscrollcommand=scrollbar.set,width=25)
        scrollbar.config(command=self.paises.yview)
        scrollbar.grid(row=0,column=4,rowspan=7,sticky=N+S)
        self.paises.grid(row=0,column=3,rowspan=7,sticky=N+S)
        for pais in self.s_Paises:
            self.paises.insert('end',pais)
            
        #LIST BOX MESES
        s_Meses = ['January','February','March','April','May','June',
                   'July','August','September','October','November','December']
        iMonth = date.today().replace(day=1)
        iMonth = iMonth - timedelta(days=1)
        iMonth = s_Meses[iMonth.month - 1]
        #DROP DOWN
        self.varMonth = StringVar(self.main)
        self.varMonth.set(iMonth)
        self.optMonth = OptionMenu(self.main,self.varMonth,*s_Meses)
        self.optMonth.config(width=40)
        self.optMonth.grid(row=3,column=1,columnspan=2,
                 padx=10,pady=10,
                 sticky=W+E)       

        #LIST BOX TYPE
        sType = ['Inbound','Compliance','Outbound','Cantonal']
        #DROP DOWN
        self.varType = StringVar(self.main)
        self.varType.set('Inbound')
        self.optType = OptionMenu(self.main,self.varType,*sType)
        self.optType.config(width=15)
        self.optType.grid(row=4,column=1,
                 padx=10,pady=10,
                 sticky=W+E)
        
        #CHECK BOX MID MONTH
        self.varMid = IntVar()
        self.MidSRR = Checkbutton(self.main,text='Mid SRR',variable=self.varMid)
        self.MidSRR.grid(row=4,column=2,
                  padx=10,pady=10,
                  sticky=W+E)

        #CHECK BOX MANUAL INPUT
        self.varManual = IntVar()
        self.Manual = Checkbutton(self.main,text='Manual Input',variable=self.varManual)
        self.Manual.grid(row=5,column=2,
                  padx=10,pady=10,
                  sticky=W+E)
        
    def change_GU(self, *args):
        self.update_TY()
        self.update_Countries()
        self.update_Year()

    def change_TY(self, *args):
        self.update_Countries()
        self.update_Year()
        
    def update_TY(self):
        #ACTUALIZA TY
        s_TY = BD.df_CountryList[(BD.df_CountryList['Geographic Unit Description']==self.varGU.get())]['Tax Year Start Date'].drop_duplicates().sort_values().reset_index(drop=True)
        menu = self.opTY['menu']
        menu.delete(0, 'end')        
        self.varTY.set(s_TY[0])         
        for TY in s_TY:
            menu.add_command(label=TY, command=lambda nation=TY: self.varTY.set(nation)) 
                
    def update_Countries(self):
        #ACTUALIZA BOX PAISES
        self.s_Paises = BD.df_CountryList[(BD.df_CountryList['Geographic Unit Description']==self.varGU.get()) &
                              (BD.df_CountryList['Tax Year Start Date']==self.varTY.get())]['Country Name'].drop_duplicates().sort_values().reset_index(drop=True)
        self.paises.delete(0,'end')
        for pais in self.s_Paises:
            self.paises.insert('end',pais)
    
    def update_Year(self):
        fecha = (datetime.strptime(self.varTY.get()[-2:] + '-' + self.varTY.get()[3:-3] + '-' + str(datetime.now().year),'%d-%B-%Y') - DateOffset(days=1))
        if fecha >= datetime.now().replace(day=1):
            self.varYear.set(datetime.now().year -1)
        else:
            self.varYear.set(datetime.now().year)
            
     
    def actionButtonPress(self):
        img = ip.get_screenshot('srr tool') # get screenshot

        print('Starting Analysis ' + datetime.now().time().strftime('%H:%M:%S'))
        #OBTENGO LA INFORMACICON SELECCIONADA
        hostCountries = [] #CREO LISTA VACIA PARA LOS HOST COUNTRIES
        seleccion = self.paises.curselection() #TOMA LOS PAISES SELECCIONADOS
        if seleccion: #SI SELECCIONO ALGUN PAIS OBTENGO LOS PAISES SELECCIONADOS Y DEJO EL GU EN BLANCO
            reportByGU = False
            for item in seleccion:
                hostCountries.append(self.paises.get(item))
        else: #SI NO SELECCIONO NINGUNO OBTIENE AL INFO DE LA SERIE DEFINIDA ANTES
            reportByGU = True
            hostCountries = self.s_Paises.tolist()        
        hostGU = self.varGU.get() #COMPLETA EL HOST GU PARA CORRER LOS REPORTES POR GU
        srrType = self.varType.get()  
        #OBTENGO EL TY SELECCIONADO   
        if self.varTY.get()[3:-3]=='January':
            tyEnd = (datetime.strptime(self.varTY.get()[-2:] + '-' + self.varTY.get()[3:-3] + '-' + str(self.varYear.get()),'%d-%B-%Y') - DateOffset(days=1)).replace(year=int(self.varYear.get()))
        else:
            tyEnd = (datetime.strptime(self.varTY.get()[-2:] + '-' + self.varTY.get()[3:-3] + '-' + str(int(self.varYear.get())+1),'%d-%B-%Y') - DateOffset(days=1))  # .replace(year=int(self.varYear.get())+1)
        
        import calendar
       
        tyStart = tyEnd + DateOffset(days=1) - DateOffset(years=1)
        if (calendar.isleap(tyStart.year)) and (tyStart.month==2):
            tyStart += DateOffset(days=1)
        #ARMA FECHAS CON LAS QUE VA A TENER QUE CORRER EL REPORTE
        reportStart = tyStart.replace(day=1,year=tyStart.year-1)

        reportEnd = datetime.strptime(str(self.varYear.get()) + self.varMonth.get() , '%Y%B') + timedelta(days=31)
        reportEnd = reportEnd.replace(day=1) - timedelta(days=1)  
#        myyear = self.varYear.get()
#        Esmid = self.varMid.get()
#        #esto es nuevo // nahuel.a.rios
#        
#        
#        import calendar
#        if calendar.isleap(int(self.varYear.get()) + 1) == True:
#           dia  = timedelta(days=4)
#           reportEnd = reportEnd - dia 
#        else:
#            reportEnd = reportEnd.replace(day=1) - timedelta(days=1)
#                
#            
            
        if tyStart > reportEnd:
            if reportEnd.day == 29:
                reportEnd = reportEnd.replace(day=reportEnd.day-1)
                reportEnd = reportEnd.replace(year=reportEnd.year+1)
            else:
                reportEnd = reportEnd.replace(year=reportEnd.year+1)
        if self.varMid.get() > 0:
            reportEnd = reportEnd.replace(day=15)
        taxInfoYear = (tyStart + DateOffset(years=1) - timedelta(days=1)).year

        #CREA CARPETA EN EL DESKTOP
        if self.varMid.get()>0:
            srrName = 'Mid '
        else:
            srrName = ''
        if reportByGU:
            srrName = srrName + 'SRR - ' + hostGU
            if srrType!='Inbound':
                srrName = srrName + ' ' + srrType + ' '
            if tyStart.month == 1:
                srrName = srrName + ' CY - '
            else:
                srrName = srrName + ' Non CY - '
        else:
            srrName = srrName + 'SRR - ' + ",".join(hostCountries) + ' - '
        srrName = srrName.replace('/', '-') + reportEnd.strftime('%B %Y')       
        deskPath = 'C:\\Users\\' + environ.get('USERNAME') + '\\Desktop\\' + srrName + ' - ' + datetime.now().strftime('%d%m%y-%H%M%S')     
        makedirs(deskPath)
        suppDoc = deskPath + '\\Supporting Doc'
        makedirs(suppDoc)

        img.save(suppDoc + '\\Input_Parameters.png', 'PNG') # save screenshot as \Supporting Doc\Input_Parameters.png
        
        if self.varManual.get()>0:
            try:
                BD.driver.quit()
            except:
                pass
            files = filedialog.askopenfilenames(title='SSRS Strategic Repatriation Report')
            df_Travel_Plan = None
            for file in files:
                print('Importing ' + str(files.index(file)+1) + ' of ' + str(len(files)) + ' ' + datetime.now().time().strftime('%H:%M:%S'))
                xlFile = ExcelFile(file)
                sheets = xlFile.sheet_names
                del sheets[0]
                for sheet in sheets:
                    if df_Travel_Plan is None:
                        df_Travel_Plan = xlFile.parse(sheet_name=sheet)
                    else: 
                        df_Travel_Plan = df_Travel_Plan.append(xlFile.parse(sheet_name=sheet))
                copyfile(file,suppDoc + '\\' + path.basename(file))
        else:
            Ch = IE_Reports(BD.driver,reportByGU,hostCountries,hostGU,tyStart,reportStart,reportEnd,taxInfoYear,suppDoc,srrType)
            df_Travel_Plan = Ch.df_Travel_Plan
        # Cambiar Home y Destination Country/Location a Country. Con Try para poder usar archivos viejos
        try:
            df_Travel_Plan.rename(columns={
                'Home Country/Location':'Home Country',
                'Destination Country/Location':'Destination Country'
            },
            inplace=True)
        except:
            pass
        
        
      
        df_Travel_Plan.rename(columns={
                'Home Market Unit':'Home Geographic Unit',
                'Destination Market Unit':'Destination Geographic Unit'},inplace=True)
        
        US_Lista = ["Northeast", "Midwest", "West" ,"South" , "United States"]
        
        df_Travel_Plan.loc[(df_Travel_Plan['Destination Geographic Unit'].isin(US_Lista)),['Destination Geographic Unit']] = "United States"
        df_Travel_Plan.loc[(df_Travel_Plan['Home Geographic Unit'].isin(US_Lista)),['Home Geographic Unit']] = "United States"

        df_Travel_Plan = df_Travel_Plan[(df_Travel_Plan['Policy Type'] != "750")]
        df_Travel_Plan = df_Travel_Plan[(df_Travel_Plan['Travel Plan Number'] != "TRA0204802")]
        df_Travel_Plan = df_Travel_Plan[(df_Travel_Plan['Current Employee ID'] != "Cristine Malapad")]
        df_Travel_Plan = df_Travel_Plan[(df_Travel_Plan['Program Name/Transfer Type'] != "Localization")]
        df_Travel_Plan = df_Travel_Plan[(df_Travel_Plan['Program Name/Transfer Type'] != "Localization Phase")]
        df_Travel_Plan = df_Travel_Plan[(df_Travel_Plan['Program Name/Transfer Type'] != "Localization Phase")]
        df_Travel_Plan = df_Travel_Plan[(df_Travel_Plan['Policy Type'] != "740") & (df_Travel_Plan['Home Country'] != df_Travel_Plan['Destination Country'])]

        
        
        #SACA PN QUE NO SON NUMERO
        df_Travel_Plan['Current Employee ID'] = to_numeric(df_Travel_Plan['Current Employee ID'],errors='coerce')
        df_Travel_Plan['Travel Plan Employee ID'] = to_numeric(df_Travel_Plan['Travel Plan Employee ID'],errors='coerce')
        df_Travel_Plan.loc[df_Travel_Plan['Current Employee ID'].isnull(),'Current Employee ID'] = df_Travel_Plan['Travel Plan Employee ID']
        df_a = df_Travel_Plan[df_Travel_Plan['Travel Plan Employee ID'].isnull()]
        for i in df_a.index.values:
            df_Travel_Plan.loc[df_Travel_Plan.index == i,'Travel Plan Employee ID'] = df_Travel_Plan.loc[df_Travel_Plan.index == i]['Current Employee ID']
        df_Travel_Plan.loc[df_Travel_Plan['Travel Plan Employee ID'].isnull(),'Travel Plan Employee ID'] = df_Travel_Plan['Current Employee ID']
        df_Travel_Plan.dropna(subset=['Current Employee ID'],inplace=True)
        df_Travel_Plan.dropna(subset=['Travel Plan Employee ID'],inplace=True)
        
        ######FORMATO DEL TRAVEL PLAN
        df_Travel_Plan.loc[df_Travel_Plan['Tax Information Year']!=taxInfoYear,
                           ['Tax Information Year','Compensation Needed','Tax Compliance (Home)','Resident (Home)',
                            'Tax Compliance (Destination)','Tax ID Number (Host)','Resident (Host)']] = None
        df_Travel_Plan.sort_values(['Travel Plan Number','Tax Information Year','Expiration Date'],ascending=[True,False,False],inplace=True)           
        df_Travel_Plan.drop_duplicates(['Travel Plan Number'],inplace=True)   
        df_Travel_Plan.reset_index(drop=True,inplace=True)   
        df_Travel_Plan['Policy Type'] = df_Travel_Plan['Policy Type'].astype(str)
#        df_Travel_Plan['Policy Type'] = to_numeric(df_Travel_Plan['Policy Type'], errors='coerce')
        df_Travel_Plan['Travel Start Date'] = to_datetime(df_Travel_Plan['Travel Start Date'], errors='coerce')

        df_Travel_Plan['Travel End Date'] = to_datetime(df_Travel_Plan['Travel End Date'], errors='coerce')
        df_Travel_Plan['Taxable Start Date'] = to_datetime(df_Travel_Plan['Taxable Start Date'], errors='coerce')
        df_Travel_Plan['Taxable End Date'] = to_datetime(df_Travel_Plan['Taxable End Date'], errors='coerce')
        
        if srrType!='Outbound':
            df_SRR = TCDL_Analysis(df_Travel_Plan,BD.df_CountryList,BD.df_TCDLTool,tyStart,hostGU)
        else:
            df_SRR = df_Travel_Plan
        
        df_SRR['SRR Comments']=None
        Col_SRR = list(df_SRR)
        
        if len(df_SRR) > 0:
            if self.varManual.get()>0:
                files = filedialog.askopenfilenames(title='Location by Day Report')
                df_LBD = None
                for file in files:
                    print('Importing ' + str(files.index(file)+1) + ' of ' + str(len(files)) + ' ' + datetime.now().time().strftime('%H:%M:%S'))
                    if df_LBD is None:
                        df_LBD = read_csv(file)
                    else: 
                        df_LBD = df_LBD.append(read_csv(file))
                    copyfile(file,suppDoc + '\\' + path.basename(file))
            else:
                df_LBD = Ch.LBD_Report(df_SRR)
                
                #######CORRER POR PN
                periods_required = ((reportEnd.year - reportStart.year) * 12 + reportEnd.month - reportStart.month + 1) * 2
                if self.varMid.get()>0:
                    periods_required += -1
                
                df_P = df_LBD[(df_LBD['Date'] == 1) | (df_LBD['Date'] == 16)].drop_duplicates(subset=['EnterpriseId','Year','Month','Date'])
                df_P['COMB'] = df_P['Date'].astype(str) + df_P['Month'] + df_P['Year'].astype(str)
                df_P = df_P[['EnterpriseId','COMB']]
                df_P = DataFrame.pivot_table(df_P,values='COMB',index=['EnterpriseId'],aggfunc=lambda x: len(x.unique())).fillna(0)
                df1 = df_SRR[['Current Employee ID','Enterprise ID']]; df1.rename(columns={'Current Employee ID':'PN','Enterprise ID':'EnterpriseId'},inplace=True)
                df2 = df_SRR[['Travel Plan Employee ID','Enterprise ID']]; df2.rename(columns={'Travel Plan Employee ID':'PN','Enterprise ID':'EnterpriseId'},inplace=True)
                df = df1.append(df2).drop_duplicates()            
                df = df.merge(df_P,how='left',left_on='EnterpriseId',right_index=True).fillna(0)
                df = df[df['COMB'] < periods_required].drop(columns=['COMB'])
                
                if len(df) > 0:
                    df_LBD = df_LBD.append(Ch.LBD_Report_by_PN(df))
            
            #METE EL PEOPLE KEY EN EL LBD
            print('Populating People Key into LBD ' + datetime.now().time().strftime('%H:%M:%S'))
            df = df_SRR[['Enterprise ID','People Key']]
            df1 = df_LBD.merge(df,how='inner',left_on='EnterpriseId',right_on='Enterprise ID').drop_duplicates()
            df = df_SRR[['Current Employee ID','People Key']]
            df2 = df_LBD.merge(df,how='inner',left_on='PersonnelNbr',right_on='Current Employee ID').drop_duplicates()
            df = df_SRR[['Travel Plan Employee ID','People Key']]
            df3 = df_LBD.merge(df,how='inner',left_on='PersonnelNbr',right_on='Travel Plan Employee ID').drop_duplicates()
            df_LBD = df1.append(df2.append(df3,sort=False),sort=False)[['People Key','HomeCountry', 'Location1','Year','Month','Date','TypeOfDay']].drop_duplicates().reset_index(drop=True)
            
            df_SRR = Tax_Threshold(BD.df_iTNAT,df_SRR,BD.df_Exceptions,srrType)
            df_SRR,df_LBD = Days_Analysis(df_SRR,df_LBD,BD.df_CountryNames,reportStart,reportEnd,tyStart)
            df_SRR = Status(df_LBD,df_SRR,reportStart,reportEnd)
            df_SRR = Control(df_SRR,df_LBD,reportStart,reportEnd)
            df_SRR = Arrange(df_SRR,Col_SRR,reportStart,reportEnd,tyStart)
            df_SRR = Percents(df_SRR,df_LBD,tyStart)
            df_SRR = DOA_Analysis(df_SRR,df_LBD,tyStart)
        
        else:
            a = []
            for i in range(24):
                a.append((reportStart + DateOffset(months=i)).strftime('%B-%Y'))
                a.append('Status ' + (reportStart + DateOffset(months=i)).strftime('%B-%Y'))
                a.append('Control ' + (reportStart + DateOffset(months=i)).strftime('%B-%Y'))
            b = ['Total Days','Tax Threshold','Days to Reach Taxability','Missing MyTE Period Within Assignnment',
                 'Days Out of Assignment (total number of days out)','0 Days Within Assignment (number of period with 0 days)',
                 'Taxability This Month','Taxability Next Month','MAXA','Host PP Days in Current TY','Total PP Days in Current TY',
                 'Percent PP Days in Current TY','Host WD Days in Current TY','Total WD Days in Current TY',
                 'Percent WD Days in Current TY','Host PP Days in Current Assignment','Total PP Days in Current Assignment',
                 'Percent PP Days in Current Assignment','Host WD Days in Current Assignment','Total WD Days in Current Assignment',
                 'Percent WD Days in Current Assignment','Days Out Before ' + reportStart.strftime('%d-%b-%Y'),
                 'Days Out Within 15 to Start Date','Days Out Within 15 to End Date','Days Out Within 30 to Start Date',
                 'Days Out Within 30 to End Date','Days Out Over 30 to Start Date','Days Out Over 30 to End Date']
            a = a + b
            df_SRR = concat([df_SRR,DataFrame(columns=a)],sort=False)
            Col_SRR = Col_SRR + a
            df_SRR = df_SRR[Col_SRR]
        
        Final_Report(df_SRR,BD.list_ARE,reportStart,reportEnd,tyStart,srrType,deskPath,srrName,hostCountries,hostGU,self.varMid.get(),self.varManual.get())
                
        print('Analysis Complete ' + datetime.now().time().strftime('%H:%M:%S'))
        messagebox.showinfo('SRR/IC10 Tool','Analysis Complete')
        
#SE TRAEN TODAS LAS LISTAS BASE NECESARIAS
BD = BaseData()

#CREO LISTADO DE GUs PARA EL DROP DOWN
main = srrGUI(BD)
main.title('SRR Tool ' + BD.version)
# main.iconbitmap('SMM.ico')
main.mainloop()

